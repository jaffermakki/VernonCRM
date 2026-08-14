import csv
import io
import json
import os
import re
import secrets
from datetime import datetime, timedelta
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent  # crm_python/

# Error monitoring: only activates if SENTRY_DSN is set (Settings → get a free
# DSN at sentry.io, add it as a Railway/Render env var). Without it, this is a
# complete no-op — local dev and any deployment that hasn't set it up yet
# behave exactly as before. traces_sample_rate is kept low since this is a
# small-shop app, not a high-traffic service — no need to pay for full tracing.
_SENTRY_DSN = os.environ.get("SENTRY_DSN", "")
if _SENTRY_DSN:
    import sentry_sdk
    sentry_sdk.init(
        dsn=_SENTRY_DSN,
        traces_sample_rate=0.1,
        environment=os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RENDER") or "local",
        send_default_pii=False,  # don't auto-attach request bodies/cookies — this app handles customer PII
    )

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, Request, Depends, Form, UploadFile, File, Query
from fastapi.responses import RedirectResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError

from .database import get_db, SessionLocal
from .models import (Staff, Product, Customer, Repair, Invoice, InvoiceLine, AuditLog, Setting, HeldCart,
                      CashSession, Supplier, PurchaseOrder, PurchaseOrderLine, RepairPart, SmsMessage,
                      TradeIn, Layaway, LayawayPayment)
from .auth import (
    hash_pin, verify_pin, is_locked, lock_seconds_remaining,
    register_pin_failure, register_pin_success, attempt_login,
    get_current_staff, role_allowed, add_audit,
)
from .tax import calc_canadian_tax, PROVINCE_LABELS
from .repairs_const import STATUS_LABELS, STATUS_ORDER, STATUS_BADGE, ISSUE_TYPES, next_status
from .barcode_gen import generate_barcode_svg
from .icons import icon_svg
from .csrf import csrf_protect, csrf_token_global, RequestContextMiddleware


def get_warranty_status(repair):
    """Warranty starts when the customer actually collects the device
    (not when work finishes), since that's when they start using it.
    Returns None if it hasn't been collected yet — nothing to check
    warranty against until then. Reads the COLLECTED entry from
    status_history rather than repair.updated_at, since updated_at can
    be touched later by unrelated edits (e.g. adjusting the final cost)
    and would silently corrupt the warranty start date."""
    if not repair.status_history:
        return None
    try:
        history = json.loads(repair.status_history)
    except (ValueError, TypeError):
        return None
    collected_entries = [h for h in history if h.get("status") == "COLLECTED" and h.get("date")]
    if not collected_entries:
        return None
    try:
        collected_date = datetime.fromisoformat(collected_entries[0]["date"])
    except (ValueError, KeyError):
        return None
    warranty_days = repair.warranty_days or 0
    expires = collected_date + timedelta(days=warranty_days)
    days_remaining = (expires - datetime.utcnow()).days
    return {
        "collected_date": collected_date, "expires": expires,
        "active": days_remaining >= 0, "days_remaining": max(0, days_remaining),
    }

from .product_const import CATEGORY_LABELS, CAT_SUBCATEGORIES
from .phone_const import PHONE_BRANDS, PHONE_MODELS, CASE_COLORS, COMMON_CASE_STYLES, LAPTOP_MODELS, CONSOLE_MODELS, COMMON_LAPTOP_GAMING_STYLES
from .notifications import send_email_receipt, send_sms, send_plain_email
from .encryption import encrypt_value, decrypt_value
from apscheduler.schedulers.background import BackgroundScheduler
from .seed import init_db

app = FastAPI(title="TechPro+ CRM")

# Use a persistent secret (set SESSION_SECRET in your environment for
# production) — a fresh random one each start would log every staff
# member out whenever the server restarts.
_SECRET_FILE = os.path.join(os.path.dirname(__file__), "..", ".session_secret")
def _get_session_secret():
    # On Railway: set SESSION_SECRET as an environment variable in the dashboard
    # On local Windows: generated once and saved to .session_secret file
    env_secret = os.environ.get("SESSION_SECRET")
    if env_secret:
        return env_secret
    if os.path.exists(_SECRET_FILE):
        with open(_SECRET_FILE) as f:
            return f.read().strip()
    new_secret = secrets.token_hex(32)
    try:
        with open(_SECRET_FILE, "w") as f:
            f.write(new_secret)
    except OSError:
        pass  # read-only filesystem (Railway) — SESSION_SECRET env var must be set
    return new_secret

# Session cookie security: on a shared shop terminal, the default 14-day
# session length is far too long — anyone approaching an unattended browser
# stays logged in for two weeks. 12 hours covers a full shift and forces a
# fresh login the next day. https_only is auto-enabled when a known hosting
# platform's env var is present (Railway/Render both set one), so local
# `uvicorn --reload` testing over plain http:// keeps working unmodified —
# set FORCE_HTTPS_COOKIES=1 explicitly for any other host that needs it.
_IS_HOSTED = bool(os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RENDER")
                  or os.environ.get("FORCE_HTTPS_COOKIES"))
app.add_middleware(
    SessionMiddleware,
    secret_key=_get_session_secret(),
    https_only=_IS_HOSTED,
    same_site="lax",
    max_age=12 * 60 * 60,
)
# Must be added after SessionMiddleware so it runs "inside" it (Starlette
# middlewares wrap outside-in in the order added, and this needs the
# session — populated by SessionMiddleware — to already be on the request
# by the time route handlers/templates read from it).
app.add_middleware(RequestContextMiddleware)

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
templates.env.globals["csrf_token"] = csrf_token_global
templates.env.globals["icon"] = icon_svg


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    """Last-resort catch for anything that isn't already handled — a bad
    query, a None slipping through, anything. Without this, FastAPI's
    default is a bare, unstyled "Internal Server Error" with no branding,
    no way back, and (depending on deployment) sometimes a raw traceback —
    not something a customer-facing till should ever show.

    Deliberately defensive: this must never itself throw, or the fallback
    for a crash becomes a worse crash. Every step below is wrapped so a
    failure in logging, in reading the session, or in rendering the error
    page still ends in *some* response rather than an unhandled 500.
    """
    import logging
    logging.getLogger("uvicorn.error").exception(
        "Unhandled exception on %s %s", request.method, request.url.path
    )
    if _SENTRY_DSN:
        try:
            import sentry_sdk
            sentry_sdk.capture_exception(exc)
        except Exception:
            pass  # error reporting failing must never block the response

    try:
        return templates.TemplateResponse(request, "error_500.html", {}, status_code=500)
    except Exception:
        # The template engine itself is having a bad day — fall back to
        # the simplest possible response rather than let this propagate.
        return HTMLResponse(
            "<h1>Something went wrong</h1><p>Please <a href='/'>go back to the dashboard</a>.</p>",
            status_code=500,
        )

# WINDOWS FIX: %-d (remove leading zero from day) is Linux-only and
# crashes on Windows with a ValueError. This custom Jinja filter does
# the same thing cross-platform by using %d and stripping manually.
def _datefmt(dt, fmt: str) -> str:
    """Cross-platform date formatting. Use {d} instead of %-d in format
    strings passed to this filter — e.g. '%B {d}, %Y'."""
    result = dt.strftime(fmt.replace("{d}", "%d"))
    result = result.replace(" 0", " ").replace("/0", "/")
    return result

templates.env.filters["datefmt"] = _datefmt

init_db()


@app.get("/healthz")
def healthz():
    """Health check for Railway/Render — deliberately does NOT touch the
    database or require login, so it stays fast and doesn't false-negative
    if the DB is briefly reconnecting (e.g. Neon waking from idle)."""
    return {"status": "ok"}



# ── helpers ──────────────────────────────────────────────────────────
def get_setting(db: Session, key: str, default=""):
    s = db.get(Setting, key)
    return s.value if s else default


def set_setting(db: Session, key: str, value: str):
    s = db.get(Setting, key)
    if s:
        s.value = value
    else:
        db.add(Setting(key=key, value=value))
    db.commit()


def _log_outgoing_sms(db: Session, phone: str, body: str, staff_name: str, customer_id: str = None):
    """Logs a successfully-sent SMS to the conversation thread. Only
    called after send_sms() reports success, so this represents messages
    the customer actually received, not failed attempts."""
    if not customer_id and phone:
        match = db.query(Customer).filter(Customer.phone == phone).first()
        customer_id = match.id if match else None
    db.add(SmsMessage(customer_id=customer_id, phone=phone, body=body, direction="out", staff_name=staff_name))
    db.commit()


def get_sms_webhook_secret(db: Session) -> str:
    """A random token embedded in the webhook URL Twilio calls when a
    customer replies. Twilio can't log in with a staff PIN, so this is
    the access control for that endpoint instead — generated once and
    reused. NOTE: this is a lighter-weight protection than Twilio's own
    request-signature validation (HMAC using your Auth Token); it stops
    randomly-guessed URLs but not a determined attacker who's somehow
    seen this exact URL. Full X-Twilio-Signature validation would be a
    further hardening step if that matters for your setup."""
    secret = get_setting(db, "sms_webhook_secret", "")
    if not secret:
        secret = secrets.token_urlsafe(24)
        set_setting(db, "sms_webhook_secret", secret)
    return secret


def cart_get(request: Request):
    return request.session.setdefault("cart", [])


def cart_totals(request: Request, db: Session):
    cart = cart_get(request)
    cart_sub = round(sum(i["price"] * i["qty"] for i in cart), 2)

    override = request.session.get("sub_override")
    sub = override if override is not None else cart_sub

    disc_mode = request.session.get("disc_mode", "$")
    disc_raw = request.session.get("disc_value", 0) or 0
    manual_disc = sub * (min(disc_raw, 100) / 100) if disc_mode == "%" else min(disc_raw, sub)

    loyalty_discount = request.session.get("loyalty_discount", 0) or 0
    store_credit_used = request.session.get("store_credit_used", 0) or 0

    disc = round(manual_disc + loyalty_discount + store_credit_used, 2)

    taxable = max(0, sub - disc)
    province = get_setting(db, "province", "ON")
    tax = calc_canadian_tax(taxable, province)
    return {
        "cart": cart, "cart_sub": cart_sub, "sub": round(sub, 2),
        "disc": disc, "disc_mode": disc_mode, "disc_raw": disc_raw,
        "loyalty_discount": loyalty_discount, "store_credit_used": store_credit_used,
        "tax": tax, "total": tax["total"],
    }


def reset_cart_overrides(request: Request):
    """Reset things that should not survive a cart-contents change —
    mirrors the original's behavior of dropping the manual subtotal
    override whenever items are added/removed/changed."""
    request.session["sub_override"] = None


def reset_customer_redemptions(request: Request):
    """Loyalty/store-credit redemptions are tied to a specific customer —
    reset them whenever the customer attached to the sale changes."""
    request.session["loyalty_discount"] = 0
    request.session["store_credit_used"] = 0


def pos_cart_context(request: Request, db: Session) -> dict:
    """Context for the POS cart/payment panel (templates/partials/pos_cart.html)
    — shared by the full /pos page render and every cart-mutating endpoint's
    AJAX response, so the panel is always built the same way no matter which
    route touched it. Deliberately leaner than the full /pos context: it
    skips the product catalog and category/brand menus, which never change
    from a cart action and would just be wasted queries on every keystroke.
    """
    customers = db.query(Customer).order_by(Customer.name).all()
    totals = cart_totals(request, db)
    customer_id = request.session.get("customer_id")
    selected_customer = db.get(Customer, customer_id) if customer_id else None
    held_carts = db.query(HeldCart).order_by(HeldCart.created_at.desc()).all()
    points_redeem_rate = float(get_setting(db, "points_redeem_rate", "100"))
    return {
        "customers": customers, "customer_id": customer_id,
        "selected_customer": selected_customer, "held_carts": held_carts,
        "points_redeem_rate": points_redeem_rate,
        "scan_error": request.session.pop("scan_error", None),
        **totals,
    }


def pos_response(request: Request, db: Session):
    """The tail call for every cart-mutating POS route. If the till's own
    fetch() JS made this request (X-Pos-Ajax header), return just the
    cart-panel HTML fragment so it can be swapped in without a page reload.
    Otherwise — no JS, a stale tab, curl, a bookmark — fall back to the
    original full-page redirect, so nothing about this behavior depends on
    JavaScript being present."""
    if request.headers.get("X-Pos-Ajax") == "1":
        return templates.TemplateResponse(request, "partials/pos_cart.html", pos_cart_context(request, db))
    return RedirectResponse("/pos", status_code=303)


def resolve_tender_split(payment_method: str, total: float, cash_part: float, card_part: float) -> tuple:
    """Decomposes a checkout's payment_method into how much of it actually
    landed as physical cash in the drawer vs. ran through the card
    terminal — the two things Cash Up has to reconcile against something
    real (a physical count, a terminal batch report). UPI, E-Transfer, and
    Store Credit settle themselves outside both of those, so they
    contribute to neither.

    A split payment is allowed to add up to more than the total (the
    checkout route only rejects it if it's short) — e.g. $5 cash + $20
    card against a $16.94 sale, expecting $8.06 change. The card side is
    exact (a terminal can't hand back partial change), so any overpayment
    comes back out of the *cash* side. The net cash this transaction
    actually leaves in the drawer is what tendered minus what went back
    out as change — which can go negative for a single transaction (this
    one nets to -$3.06: the drawer needs $3.06 more than it took in just
    to make this sale's change), even though the day's total obviously
    can't.
    """
    if payment_method == "Cash":
        return round(total, 2), 0.0
    if payment_method.startswith("Split ("):
        change = max(0.0, round(cash_part + card_part - total, 2))
        net_cash = round(cash_part - change, 2)
        return net_cash, round(card_part, 2)
    if payment_method == "Card":
        return 0.0, round(total, 2)
    return 0.0, 0.0  # UPI, E-Transfer, Store Credit — nothing to reconcile physically


def payment_method_label(payment_method: str) -> str:
    """Collapses every distinct 'Split (Cash $x + Card $y)' string down to
    one label for grouping/display — each split sale otherwise carries its
    own exact dollar amounts baked into the string, so a raw group-by would
    produce one row per split sale instead of a single 'Split' bucket."""
    if payment_method.startswith("Split ("):
        return "Split (Cash + Card)"
    return payment_method


def next_invoice_number(db: Session):
    prefix = get_setting(db, "invoice_prefix", "INV")
    counter = int(get_setting(db, "invoice_counter", "1000"))
    set_setting(db, "invoice_counter", str(counter + 1))
    return f"{prefix}-{counter}"


def next_layaway_number(db: Session):
    prefix = get_setting(db, "layaway_prefix", "LAY")
    counter = int(get_setting(db, "layaway_counter", "1000"))
    set_setting(db, "layaway_counter", str(counter + 1))
    return f"{prefix}-{counter}"


def require_login(request: Request, db: Session):
    staff = get_current_staff(request, db)
    return staff


# ── SMS WEBHOOK (incoming replies from Twilio) ──────────────────────────
# Deliberately outside any require_login check — Twilio calls this
# directly, not a logged-in staff member. Protected by the secret token
# in the URL path instead (see get_sms_webhook_secret's docstring for
# what this does and doesn't protect against).
@app.post("/webhooks/sms-reply/{secret}")
async def sms_reply_webhook(request: Request, secret: str, db: Session = Depends(get_db)):
    expected = get_sms_webhook_secret(db)
    if secret != expected:
        return Response(status_code=403)

    form = await request.form()
    from_phone = form.get("From", "")
    body = form.get("Body", "")
    if from_phone and body:
        customer = db.query(Customer).filter(Customer.phone == from_phone).first()
        db.add(SmsMessage(customer_id=customer.id if customer else None, phone=from_phone,
                           body=body, direction="in"))
        db.commit()

    # Twilio expects a TwiML (XML) response acknowledging receipt. An
    # empty <Response> means "received, no auto-reply sent" — exactly
    # what we want, since replies get read by staff, not auto-answered.
    return Response(content='<?xml version="1.0" encoding="UTF-8"?><Response></Response>',
                     media_type="application/xml")


# ── PUBLIC REPAIR STATUS LOOKUP ─────────────────────────────────────
# Deliberately unauthenticated — meant to be shared with a customer
# (e.g. "check your repair status at [shop]/status") so they don't have
# to call in. Requires BOTH the exact ticket number AND the last 4
# digits of the phone number on file to match before showing anything,
# and even on a match only exposes status-relevant fields (device,
# status, promised-by date) — never cost, customer name, or any other
# repair. Wrong/missing ticket and wrong phone digits return the exact
# same generic error, so this can't be used to enumerate which ticket
# numbers exist.
def _phone_last4(phone: str) -> str:
    digits = re.sub(r"\D", "", phone or "")
    return digits[-4:] if len(digits) >= 4 else digits


@app.get("/status", response_class=HTMLResponse)
def public_repair_status(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse(request, "public_status.html", {
        "result": None, "error": None, "shop_name": get_setting(db, "shop_name", "TechPro+"),
    })


@app.post("/status", response_class=HTMLResponse)
def public_repair_status_lookup(request: Request, ticket_no: str = Form(...), phone_last4: str = Form(...),
                                 db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    shop_name = get_setting(db, "shop_name", "TechPro+")
    result, error = None, None

    try:
        ticket_int = int(re.sub(r"\D", "", ticket_no or ""))
    except ValueError:
        ticket_int = None

    digits = re.sub(r"\D", "", phone_last4 or "")
    repair = db.query(Repair).filter(Repair.ticket_no == ticket_int).first() if ticket_int else None

    if repair and repair.customer and len(digits) == 4 and _phone_last4(repair.customer.phone) == digits:
        result = {
            "ticket_no": repair.ticket_no,
            "device": repair.device,
            "status": STATUS_LABELS.get(repair.status, repair.status),
            "status_badge": STATUS_BADGE.get(repair.status, "badge-gray"),
            "promised_by": repair.promised_by,
            "ready": repair.status in ("READY", "COMPLETED"),
        }
    else:
        error = "We couldn't find a matching ticket. Double-check your ticket number and the last 4 digits of the phone number on file."

    return templates.TemplateResponse(request, "public_status.html", {
        "result": result, "error": error, "shop_name": shop_name,
    })


# ── LOGIN ────────────────────────────────────────────────────────────
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, db: Session = Depends(get_db)):
    locked = is_locked(db)
    return templates.TemplateResponse(request, "login.html", {
        "locked": locked,
        "lock_seconds": lock_seconds_remaining(db) if locked else 0,
        "error": request.session.pop("login_error", None),
    })


@app.post("/login")
def login_submit(request: Request, pin: str = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if is_locked(db):
        request.session["login_error"] = f"Locked. Try again in {lock_seconds_remaining(db)}s."
        return RedirectResponse("/login", status_code=303)

    staff = attempt_login(db, pin)
    if staff:
        register_pin_success(db)
        request.session["staff_id"] = staff.id
        request.session["last_activity"] = datetime.utcnow().isoformat()
        add_audit(db, staff, "LOGIN", f"Staff login: {staff.name}")
        return RedirectResponse("/", status_code=303)
    else:
        register_pin_failure(db)
        request.session["login_error"] = "Incorrect PIN."
        return RedirectResponse("/login", status_code=303)


@app.get("/logout")
def logout(request: Request, db: Session = Depends(get_db)):
    staff = get_current_staff(request, db)
    if staff:
        add_audit(db, staff, "LOGOUT", f"Staff logout: {staff.name}")
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


# ── FORGOT PIN (Owner recovery only) ────────────────────────────────────
# Scoped deliberately narrow: a non-owner staff member who forgets their PIN
# can always be reset by an Owner/Manager via /staff while logged in. The
# one scenario that can lock the *entire shop* out is the Owner forgetting
# their own PIN with nobody else able to reach /staff at all — that's the
# only gap this recovers. It shares the same global lockout counter as the
# main login PIN pad, since this is a new unauthenticated entry point and a
# security-question answer can be an easier brute-force target than a PIN.
@app.get("/forgot-pin", response_class=HTMLResponse)
def forgot_pin_page(request: Request, db: Session = Depends(get_db)):
    locked = is_locked(db)
    security_question = get_setting(db, "security_question", "")
    security_configured = bool(get_setting(db, "security_answer_hash", ""))
    owners = db.query(Staff).filter(Staff.role == "owner", Staff.active == True).all()  # noqa: E712
    return templates.TemplateResponse(request, "forgot_pin.html", {
        "locked": locked, "lock_seconds": lock_seconds_remaining(db) if locked else 0,
        "security_question": security_question, "security_configured": security_configured,
        "owners": owners, "error": request.session.pop("recovery_error", None),
    })


@app.post("/forgot-pin")
def forgot_pin_submit(request: Request, staff_id: str = Form(...), security_answer: str = Form(...),
                       new_pin: str = Form(...), confirm_pin: str = Form(...),
                       db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if is_locked(db):
        request.session["recovery_error"] = f"Locked. Try again in {lock_seconds_remaining(db)}s."
        return RedirectResponse("/forgot-pin", status_code=303)

    answer_hash = get_setting(db, "security_answer_hash", "")
    if not answer_hash:
        request.session["recovery_error"] = "Account recovery isn't set up for this shop yet."
        return RedirectResponse("/forgot-pin", status_code=303)

    target = db.get(Staff, staff_id)
    if not target or target.role != "owner" or not target.active:
        request.session["recovery_error"] = "That account can't be recovered this way."
        return RedirectResponse("/forgot-pin", status_code=303)

    if not verify_pin(security_answer.strip().lower(), answer_hash):
        register_pin_failure(db)
        add_audit(db, None, "SECURITY_BLOCK", f"Wrong security answer on Forgot PIN for: {target.name}")
        db.commit()
        request.session["recovery_error"] = "Incorrect answer."
        return RedirectResponse("/forgot-pin", status_code=303)

    if not (new_pin.isdigit() and len(new_pin) == 4):
        request.session["recovery_error"] = "New PIN must be exactly 4 digits."
        return RedirectResponse("/forgot-pin", status_code=303)
    if new_pin != confirm_pin:
        request.session["recovery_error"] = "PINs don't match."
        return RedirectResponse("/forgot-pin", status_code=303)

    register_pin_success(db)
    target.pin_hash = hash_pin(new_pin)
    add_audit(db, None, "PIN_CHANGE", f"PIN reset via Forgot PIN recovery for: {target.name}")
    db.commit()
    request.session["login_error"] = f"PIN reset for {target.name}. Log in with the new PIN."
    return RedirectResponse("/login", status_code=303)


# ── DASHBOARD ────────────────────────────────────────────────────────
def _sparkline_points(values, w=100, h=28):
    """Turn a list of numbers into an SVG polyline points string, scaled
    to fit a w×h box. Flat series (all-equal or single value) render as
    a flat mid-line rather than dividing by zero."""
    if not values:
        return ""
    lo, hi = min(values), max(values)
    rng = (hi - lo) or 1
    n = len(values)
    step = w / max(n - 1, 1)
    pts = []
    for i, v in enumerate(values):
        x = round(i * step, 1)
        y = round(h - ((v - lo) / rng) * h, 1) if n > 1 else h / 2
        pts.append(f"{x},{y}")
    return " ".join(pts)


def _range_bounds(range_key: str, today):
    """Returns (start, end, prev_start, prev_end) dates for the chosen
    dashboard range, where prev_* is the immediately preceding period of
    equal length — used for the 'vs previous period' comparison."""
    if range_key == "yesterday":
        start = end = today - timedelta(days=1)
        prev_start = prev_end = today - timedelta(days=2)
    elif range_key == "7d":
        start, end = today - timedelta(days=6), today
        prev_start, prev_end = start - timedelta(days=7), start - timedelta(days=1)
    elif range_key == "30d":
        start, end = today - timedelta(days=29), today
        prev_start, prev_end = start - timedelta(days=30), start - timedelta(days=1)
    else:  # "today" (default)
        range_key = "today"
        start = end = today
        prev_start = prev_end = today - timedelta(days=1)
    return start, end, prev_start, prev_end


RANGE_LABELS = {"today": "Today", "yesterday": "Yesterday", "7d": "Last 7 days", "30d": "Last 30 days"}


@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db), date_range: str = Query("today", alias="range")):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if date_range not in RANGE_LABELS:
        date_range = "today"

    today = datetime.utcnow().date()
    all_invoices = db.query(Invoice).all()
    start, end, prev_start, prev_end = _range_bounds(date_range, today)

    in_range = [i for i in all_invoices if start <= i.date.date() <= end]
    in_prev = [i for i in all_invoices if prev_start <= i.date.date() <= prev_end]

    sales_total = round(sum(i.total for i in in_range), 2)
    sales_prev = round(sum(i.total for i in in_prev), 2)
    count_total = len(in_range)
    count_prev = len(in_prev)

    def _trend_pct(cur, prev):
        if prev > 0:
            return round((cur - prev) / prev * 100, 1)
        return 100.0 if cur > 0 else 0.0

    sales_trend_pct = _trend_pct(sales_total, sales_prev)
    count_trend_pct = _trend_pct(count_total, count_prev)

    # Daily series across the selected range, oldest first — powers both
    # the big 7-day chart (kept fixed at 7 days regardless of the range
    # picker, since it's a stable at-a-glance chart) and the sparklines
    # inside the Sales/Invoices cards (which follow the selected range,
    # capped at 30 points so a 30-day sparkline stays readable).
    span_days = (end - start).days + 1
    daily_sales, daily_counts = [], []
    for i in range(span_days - 1, -1, -1):
        d = end - timedelta(days=i)
        day_invoices = [inv for inv in all_invoices if inv.date.date() == d]
        daily_sales.append(round(sum(inv.total for inv in day_invoices), 2))
        daily_counts.append(len(day_invoices))

    sales_spark = _sparkline_points(daily_sales)
    count_spark = _sparkline_points(daily_counts)

    trend_days, trend_totals = [], []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        day_total = round(sum(inv.total for inv in all_invoices if inv.date.date() == d), 2)
        trend_days.append(d.strftime("%a"))
        trend_totals.append(day_total)

    recent_invoices = db.query(Invoice).order_by(Invoice.date.desc()).limit(8).all()
    low_stock = db.query(Product).filter(Product.stock <= Product.reorder_threshold).all()
    open_repairs_count = db.query(Repair).filter(Repair.status.notin_(["COMPLETED", "COLLECTED"])).count()

    backup_warning = None
    checklist = None
    if role_allowed(staff, "owner"):
        last_backup = get_setting(db, "last_backup", "")
        total_customers = db.query(Customer).count()
        total_invoices = db.query(Invoice).count()
        if not last_backup:
            if total_invoices > 0 or total_customers > 2:
                backup_warning = "You have live data but have never exported a backup."
        else:
            days_since = (datetime.utcnow() - datetime.fromisoformat(last_backup)).days
            if days_since >= 7:
                backup_warning = f"Last backup was {days_since} day{'s' if days_since != 1 else ''} ago."

        if not get_setting(db, "hide_setup_checklist", ""):
            items = [
                {"label": "Confirm your province & tax rate", "done": db.get(Setting, "province") is not None, "href": "/settings"},
                {"label": "Invite your team", "done": db.query(Staff).count() > 1, "href": "/staff"},
                {"label": "Set up email receipts", "done": bool(get_setting(db, "smtp_host", "")), "href": "/settings"},
                {"label": "Ring up your first sale", "done": total_invoices > 0, "href": "/pos"},
                {"label": "Export a backup", "done": bool(last_backup), "href": "/export/backup"},
            ]
            done_count = sum(1 for it in items if it["done"])
            if done_count < len(items):
                checklist = {"entries": items, "done_count": done_count, "total": len(items)}

    activity = (
        db.query(AuditLog)
        .order_by(AuditLog.ts.desc())
        .limit(8)
        .all()
    )

    return templates.TemplateResponse(request, "dashboard.html", {
        "staff": staff,
        "range": date_range, "range_label": RANGE_LABELS[date_range], "range_labels": RANGE_LABELS,
        "sales_total": sales_total, "count_total": count_total,
        "sales_trend_pct": sales_trend_pct, "count_trend_pct": count_trend_pct,
        "sales_spark": sales_spark, "count_spark": count_spark,
        "trend_days": trend_days, "trend_totals": trend_totals,
        "open_repairs_count": open_repairs_count,
        "recent_invoices": recent_invoices, "low_stock": low_stock,
        "shop_name": get_setting(db, "shop_name", "TechPro+"),
        "backup_warning": backup_warning,
        "checklist": checklist,
        "activity": activity,
    })


@app.post("/dashboard/dismiss-checklist")
def dismiss_checklist(request: Request, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff or not role_allowed(staff, "owner"):
        return RedirectResponse("/", status_code=303)
    set_setting(db, "hide_setup_checklist", "1")
    return RedirectResponse("/", status_code=303)


@app.get("/api/notifications")
def api_notifications(request: Request, db: Session = Depends(get_db)):
    """Feeds the notification bell dropdown in the top bar. Lightweight
    and read-only so it can be polled from any page via fetch()."""
    staff = require_login(request, db)
    if not staff:
        return {"notifications": []}

    notes = []
    low_stock = db.query(Product).filter(Product.stock <= Product.reorder_threshold).all()
    if low_stock:
        notes.append({
            "type": "warning", "icon": "📦",
            "message": f"{len(low_stock)} product{'s' if len(low_stock) != 1 else ''} low on stock",
            "href": "/products/reorder",
        })

    overdue_cutoff = datetime.utcnow() - timedelta(days=5)
    overdue_repairs = db.query(Repair).filter(
        Repair.status.notin_(["COMPLETED", "COLLECTED"]),
        Repair.created_at < overdue_cutoff,
    ).count()
    if overdue_repairs:
        notes.append({
            "type": "danger", "icon": "🔧",
            "message": f"{overdue_repairs} repair{'s' if overdue_repairs != 1 else ''} open 5+ days",
            "href": "/repairs",
        })

    recent_replies = db.query(SmsMessage).filter(
        SmsMessage.direction == "in", SmsMessage.created_at >= datetime.utcnow() - timedelta(hours=24),
    ).count()
    if recent_replies:
        notes.append({
            "type": "warning", "icon": "💬",
            "message": f"{recent_replies} SMS repl{'ies' if recent_replies != 1 else 'y'} in the last 24h",
            "href": "/customers",
        })

    active_layaways = db.query(Layaway).filter(Layaway.status == "active").all()
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    overdue_layaways = [l for l in active_layaways if l.due_date and l.due_date < today_str]
    due_soon_layaways = [
        l for l in active_layaways
        if l.due_date and today_str <= l.due_date <= (datetime.utcnow() + timedelta(days=3)).strftime("%Y-%m-%d")
    ]
    if overdue_layaways:
        notes.append({
            "type": "danger", "icon": "📦",
            "message": f"{len(overdue_layaways)} layaway{'s' if len(overdue_layaways) != 1 else ''} past due date",
            "href": "/layaway",
        })
    elif due_soon_layaways:
        notes.append({
            "type": "warning", "icon": "📦",
            "message": f"{len(due_soon_layaways)} layaway{'s' if len(due_soon_layaways) != 1 else ''} due within 3 days",
            "href": "/layaway",
        })

    if role_allowed(staff, "owner"):
        last_backup = get_setting(db, "last_backup", "")
        if not last_backup:
            total_invoices = db.query(Invoice).count()
            if total_invoices > 0:
                notes.append({"type": "warning", "icon": "💾", "message": "No backup has ever been exported", "href": "/export/backup"})
        else:
            days_since = (datetime.utcnow() - datetime.fromisoformat(last_backup)).days
            if days_since >= 7:
                notes.append({"type": "warning", "icon": "💾", "message": f"Last backup was {days_since} days ago", "href": "/export/backup"})

    return {"notifications": notes}


@app.get("/search", response_class=HTMLResponse)
def search(request: Request, db: Session = Depends(get_db), q: str = ""):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)

    customers, invoices, repairs = [], [], []
    q = q.strip()
    if q:
        like = f"%{q}%"
        customers = db.query(Customer).filter(
            (Customer.name.ilike(like)) | (Customer.phone.ilike(like)) | (Customer.email.ilike(like))
        ).limit(15).all()
        invoices = db.query(Invoice).filter(Invoice.number.ilike(like)).limit(15).all()
        repair_filters = [Repair.device.ilike(like), Repair.issue.ilike(like), Repair.imei.ilike(like)]
        if q.isdigit():
            repair_filters.append(Repair.ticket_no == int(q))
        repairs = db.query(Repair).filter(or_(*repair_filters)).limit(15).all()

        # An IMEI/serial search should also surface any invoice a device
        # was actually sold on — not just repair tickets — since that's
        # the other half of "trace this device's history" (warranty
        # claims, stolen-phone lookups).
        if len(q) >= 4:
            imei_invoice_ids = [row[0] for row in db.query(InvoiceLine.invoice_id)
                                 .filter(InvoiceLine.imei.ilike(like)).limit(15).all()]
            if imei_invoice_ids:
                extra = db.query(Invoice).filter(Invoice.id.in_(imei_invoice_ids)).all()
                seen = {i.id for i in invoices}
                invoices += [i for i in extra if i.id not in seen]

    return templates.TemplateResponse(request, "search.html", {
        "staff": staff, "q": q,
        "customers": customers, "invoices": invoices, "repairs": repairs,
        "shop_name": get_setting(db, "shop_name", "TechPro+"),
    })




# ── POS ──────────────────────────────────────────────────────────────
@app.get("/pos", response_class=HTMLResponse)
def pos_page(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    products = db.query(Product).all()
    all_brands = sorted({p.subcategory for p in products if p.subcategory})
    repair_ctx_id = request.session.get("pos_repair_id")
    repair_ctx = db.get(Repair, repair_ctx_id) if repair_ctx_id else None
    product_groups, ungrouped_products = group_products_by_variant(products)
    variant_group_data = {
        g["name"]: [
            {"pid": v.id, "name": v.name, "sku": v.sku, "price": v.price, "stock": v.stock,
             "reorder_threshold": v.reorder_threshold, "brand": v.subcategory or "",
             "phone_brand": v.phone_brand or "", "phone_model": v.phone_model or "", "color": v.color or ""}
            for v in g["variants"]
        ]
        for g in product_groups
    }
    phone_menu = {}
    for p in products:
        if p.phone_brand and p.phone_model:
            phone_menu.setdefault(p.phone_brand, set()).add(p.phone_model)
    phone_menu = {brand: sorted(models) for brand, models in phone_menu.items()}
    all_phone_brands = sorted(phone_menu.keys())
    return templates.TemplateResponse(request, "pos.html", {
        "staff": staff, "products": products,
        "product_groups": product_groups, "ungrouped_products": ungrouped_products,
        "variant_group_data": variant_group_data,
        "phone_menu": phone_menu, "all_phone_brands": all_phone_brands,
        "category_labels": CATEGORY_LABELS, "all_brands": all_brands,
        "cat_subcategories": CAT_SUBCATEGORIES,
        "repair_ctx": repair_ctx,
        **pos_cart_context(request, db),
    })


@app.post("/pos/scan")
def pos_scan(request: Request, sku: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    """Looks up a scanned/typed SKU and hands it to the same price-check
    modal a tapped product tile already uses — instead of adding straight
    to the cart, so staff get one chance to confirm the price or apply an
    on-the-spot discount before it's in the sale. Every barcode is
    inherently unambiguous (matches exactly one Product row), so this
    never needs the variant-picker modal a grouped tile can trigger.

    Only the AJAX (JS-enabled) path behaves this way. A non-AJAX request
    — JS failed, a stale tab, curl — has no way to show a modal at all,
    so it falls back to the original direct-add-to-cart behavior, same as
    every other cart-mutating route's no-JS safety net.
    """
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    needle = sku.strip().upper()
    is_ajax = request.headers.get("X-Pos-Ajax") == "1"

    if not needle:
        if is_ajax:
            return {"found": False, "error": ""}
        return pos_response(request, db)

    product = db.query(Product).filter(Product.sku.ilike(needle)).first()
    if not product:
        product = db.query(Product).filter(Product.name.ilike(f"%{needle}%")).first()

    if not product:
        if is_ajax:
            return {"found": False, "error": f'No product found for "{sku}"'}
        request.session["scan_error"] = f'No product found for "{sku}"'
        return pos_response(request, db)

    if is_ajax:
        return {"found": True, "product": {
            "pid": product.id, "name": product.name, "sku": product.sku,
            "price": product.price, "stock": product.stock,
        }}

    cart = cart_get(request)
    for item in cart:
        if item["product_id"] == product.id:
            item["qty"] += 1
            break
    else:
        cart.append({"product_id": product.id, "name": product.name, "sku": product.sku, "price": product.price, "qty": 1})
    request.session["cart"] = cart
    request.session["sub_override"] = None
    return pos_response(request, db)


@app.post("/pos/add/{product_id}")
def pos_add(request: Request, product_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    product = db.get(Product, product_id)
    if product:
        cart = cart_get(request)
        for item in cart:
            if item["product_id"] == product_id and item["price"] == product.price:
                item["qty"] += 1
                break
        else:
            cart.append({"product_id": product.id, "name": product.name, "sku": product.sku, "price": product.price, "qty": 1})
        request.session["cart"] = cart
        request.session["sub_override"] = None
    return pos_response(request, db)


@app.post("/pos/add-custom")
def pos_add_custom(request: Request, product_id: str = Form(...),
                    custom_price: float = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    """Add a product to the cart at a user-specified price instead of the stored price.
    Each custom-price entry is always a separate line item (never merged with other
    entries for the same product, since the price may differ)."""
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    product = db.get(Product, product_id)
    if product and custom_price >= 0:
        cart = cart_get(request)
        cart.append({
            "product_id": product.id,
            "name": product.name,
            "sku": product.sku,
            "price": round(custom_price, 2),
            "qty": 1,
        })
        request.session["cart"] = cart
        request.session["sub_override"] = None
    return pos_response(request, db)


@app.post("/pos/qty/{idx}")
def pos_qty(request: Request, idx: int, qty: int = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    cart = cart_get(request)
    if 0 <= idx < len(cart):
        if qty <= 0:
            cart.pop(idx)
        else:
            cart[idx]["qty"] = qty
        request.session["cart"] = cart
        request.session["sub_override"] = None
    return pos_response(request, db)


@app.post("/pos/imei/{idx}")
def pos_imei(request: Request, idx: int, imei: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    """Attaches an IMEI/serial to a specific cart line before checkout —
    copied onto the InvoiceLine at checkout time so a sold device's
    serial is on permanent record against that sale."""
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    cart = cart_get(request)
    if 0 <= idx < len(cart):
        cart[idx]["imei"] = imei.strip()
        request.session["cart"] = cart
    return pos_response(request, db)


@app.post("/pos/remove/{idx}")
def pos_remove(request: Request, idx: int, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    cart = cart_get(request)
    if 0 <= idx < len(cart):
        cart.pop(idx)
        request.session["cart"] = cart
        request.session["sub_override"] = None
    return pos_response(request, db)


@app.post("/pos/clear")
def pos_clear(request: Request, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    request.session["cart"] = []
    request.session["sub_override"] = None
    request.session["disc_value"] = 0
    request.session["disc_mode"] = "$"
    request.session["customer_id"] = None
    request.session["pos_repair_id"] = None
    reset_customer_redemptions(request)
    return pos_response(request, db)


@app.post("/pos/subtotal")
def pos_subtotal(request: Request, value: float = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    request.session["sub_override"] = max(0, value)
    return pos_response(request, db)


@app.post("/pos/discount")
def pos_discount(request: Request, mode: str = Form(...), value: float = Form(0), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    request.session["disc_mode"] = mode if mode in ("$", "%") else "$"
    request.session["disc_value"] = max(0, value)
    return pos_response(request, db)


@app.post("/pos/customer")
def pos_customer(request: Request, customer_id: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    request.session["customer_id"] = customer_id or None
    reset_customer_redemptions(request)  # redemptions are tied to whoever was previously attached
    return pos_response(request, db)


@app.post("/pos/redeem-points")
def pos_redeem_points(request: Request, points: int = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    customer_id = request.session.get("customer_id")
    customer = db.get(Customer, customer_id) if customer_id else None
    rate = float(get_setting(db, "points_redeem_rate", "100"))
    if customer:
        # Round down to a whole multiple of the redemption rate, same as the original
        pts = (points // int(rate)) * int(rate)
        if 0 < pts <= (customer.points or 0):
            dollar_value = pts / rate
            request.session["loyalty_discount"] = (request.session.get("loyalty_discount", 0) or 0) + dollar_value
    return pos_response(request, db)


@app.post("/pos/redeem-credit")
def pos_redeem_credit(request: Request, amount: float = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    customer_id = request.session.get("customer_id")
    customer = db.get(Customer, customer_id) if customer_id else None
    if customer and 0 < amount <= (customer.store_credit or 0):
        request.session["store_credit_used"] = (request.session.get("store_credit_used", 0) or 0) + amount
    return pos_response(request, db)


@app.post("/pos/hold")
def pos_hold(request: Request, name: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    cart = cart_get(request)
    if not cart:
        return pos_response(request, db)
    held = HeldCart(
        name=name.strip() or f"Hold {datetime.utcnow().strftime('%H:%M:%S')}",
        cart_json=json.dumps(cart),
        customer_id=request.session.get("customer_id"),
        disc_mode=request.session.get("disc_mode", "$"),
        disc_value=request.session.get("disc_value", 0) or 0,
    )
    db.add(held)
    add_audit(db, staff, "HOLD_CART", f"Cart held: {held.name}")
    db.commit()

    request.session["cart"] = []
    request.session["sub_override"] = None
    request.session["disc_value"] = 0
    request.session["disc_mode"] = "$"
    request.session["customer_id"] = None
    reset_customer_redemptions(request)
    return pos_response(request, db)


@app.post("/pos/recall/{held_id}")
def pos_recall(request: Request, held_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    held = db.get(HeldCart, held_id)
    if held:
        request.session["cart"] = json.loads(held.cart_json)
        request.session["sub_override"] = None
        request.session["customer_id"] = held.customer_id
        request.session["disc_mode"] = held.disc_mode
        request.session["disc_value"] = held.disc_value
        reset_customer_redemptions(request)
        add_audit(db, staff, "RECALL_CART", f"Cart recalled: {held.name}")
        db.delete(held)
        db.commit()
    return pos_response(request, db)


@app.post("/pos/held/{held_id}/delete")
def pos_held_delete(request: Request, held_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    if not require_login(request, db):
        return RedirectResponse("/login", status_code=303)
    held = db.get(HeldCart, held_id)
    if held:
        db.delete(held)
        db.commit()
    return pos_response(request, db)


@app.post("/pos/checkout")
def pos_checkout(request: Request, payment_method: str = Form("Cash"),
                  tendered: float = Form(0),
                  cash_part: float = Form(0),
                  card_part: float = Form(0),
                  card_reference: str = Form(""),
                  db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)

    cart = cart_get(request)
    if not cart:
        return RedirectResponse("/pos", status_code=303)

    totals = cart_totals(request, db)
    total = totals["total"]

    # ── Payment resolution ──────────────────────────────────────────
    # Split payment: both cash_part and card_part provided
    if cash_part > 0 and card_part > 0:
        split_total = round(cash_part + card_part, 2)
        if split_total < total:
            return HTMLResponse(f"Split payment total (${split_total:.2f}) is less than the invoice total (${total:.2f}).", status_code=400)
        payment_method = f"Split (Cash ${cash_part:.2f} + Card ${card_part:.2f})"
        tendered = cash_part
        change_given = max(0, round(split_total - total, 2))
    elif payment_method == "Cash" and tendered > 0 and tendered < total:
        return HTMLResponse(f"Tendered amount (${tendered:.2f}) is less than the total (${total:.2f}).", status_code=400)
    else:
        change_given = max(0, round(tendered - total, 2)) if tendered > 0 else 0

    customer_id = request.session.get("customer_id")
    customer = db.get(Customer, customer_id) if customer_id else None

    loyalty_discount = totals["loyalty_discount"]
    store_credit_used = totals["store_credit_used"]
    rate = float(get_setting(db, "points_redeem_rate", "100"))
    loyalty_pts_used = round(loyalty_discount * rate)

    repair_ctx_id = request.session.get("pos_repair_id")
    repair_ctx = db.get(Repair, repair_ctx_id) if repair_ctx_id else None

    cash_amount, card_amount = resolve_tender_split(payment_method, total, cash_part, card_part)
    card_reference = card_reference.strip()[:64]  # a reasonable cap; this is a receipt code, not free text

    # Two checkouts completing at the exact same instant (two terminals, two
    # staff) could theoretically read the same counter value before either
    # writes it back. The database's unique constraint on `number` guarantees
    # a duplicate can never actually be saved — this retry loop just makes
    # sure that collision produces a fresh number instead of a failed sale.
    # A fresh Invoice object is built each attempt rather than reusing one
    # across a rollback, to avoid relying on SQLAlchemy's session-state
    # behavior for an object involved in a failed flush — this is money, so
    # the extra caution here is deliberate.
    invoice = None
    for attempt in range(5):
        candidate = Invoice(
            number=next_invoice_number(db),
            customer_id=customer.id if customer else None,
            staff_id=staff.id,
            repair_id=repair_ctx.id if repair_ctx else None,
            payment_method=payment_method,
            cash_amount=cash_amount,
            card_amount=card_amount,
            card_reference=card_reference,
            subtotal=totals["sub"],
            discount=totals["disc"],
            loyalty_pts_used=loyalty_pts_used,
            store_credit_used=store_credit_used,
            tendered=tendered,
            change_given=change_given,
            tax_breakdown=json.dumps(totals["tax"]["lines"]),
            tax_total=totals["tax"]["tax_total"],
            total=total,
        )
        db.add(candidate)
        try:
            db.flush()
            invoice = candidate
            break
        except IntegrityError:
            db.rollback()
            if attempt == 4:
                raise
    if invoice is None:
        raise RuntimeError("Could not generate a unique invoice number after 5 attempts")

    for item in cart:
        db.add(InvoiceLine(invoice_id=invoice.id, product_id=item["product_id"],
                            name=item["name"], sku=item.get("sku", ""), qty=item["qty"], price=item["price"],
                            imei=item.get("imei", "")))
        product = db.get(Product, item["product_id"]) if item["product_id"] else None
        if product:
            product.stock = max(0, product.stock - item["qty"])

    if customer:
        # Spend redemptions first (mirrors the original's order of operations)
        if store_credit_used > 0:
            customer.store_credit = round((customer.store_credit or 0) - store_credit_used, 2)
            add_audit(db, staff, "STORE_CREDIT", f"Redeemed ${store_credit_used:.2f} store credit for {customer.name}")
        if loyalty_pts_used > 0:
            customer.points = max(0, (customer.points or 0) - loyalty_pts_used)
            add_audit(db, staff, "LOYALTY", f"Redeemed {loyalty_pts_used} points for ${loyalty_discount:.2f} — {customer.name}")
        # Earn new points on the final total
        points_per_dollar = float(get_setting(db, "points_per_dollar", "1"))
        customer.points = (customer.points or 0) + int(total * points_per_dollar)
        customer.spent = round((customer.spent or 0) + total, 2)
        customer.last_visit = datetime.utcnow().isoformat()

    add_audit(db, staff, "INVOICE_CREATE", f"Invoice {invoice.number} — ${total:.2f} — {payment_method}")

    collected_note = ""
    if repair_ctx and repair_ctx.status not in ("COLLECTED",):
        repair_ctx.status = "COLLECTED"
        repair_ctx.final_cost = total if repair_ctx.final_cost is None else repair_ctx.final_cost
        repair_ctx.updated_at = datetime.utcnow()
        rhistory = json.loads(repair_ctx.status_history) if repair_ctx.status_history else []
        rhistory.append({"status": "COLLECTED", "note": f"Paid via {payment_method}, Invoice {invoice.number}",
                          "date": datetime.utcnow().isoformat()})
        repair_ctx.status_history = json.dumps(rhistory)
        add_audit(db, staff, "REPAIR_STATUS", f"#{repair_ctx.ticket_no} → Collected (Invoice {invoice.number})")
        collected_note = f" Repair ticket #{repair_ctx.ticket_no} marked Collected."

    db.commit()

    request.session["cart"] = []
    request.session["sub_override"] = None
    request.session["disc_value"] = 0
    request.session["disc_mode"] = "$"
    request.session["customer_id"] = None
    request.session["pos_repair_id"] = None
    reset_customer_redemptions(request)
    request.session["flash"] = ("green", f"✓ Sale complete — Invoice {invoice.number} for ${total:.2f}."
                                          f"{collected_note} Use the buttons below to email or print the receipt.")
    return RedirectResponse(f"/invoices/{invoice.id}", status_code=303)


# ── INVOICES ─────────────────────────────────────────────────────────
@app.get("/invoices", response_class=HTMLResponse)
def invoices_list(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    invoices = db.query(Invoice).order_by(Invoice.date.desc()).limit(100).all()
    today = datetime.utcnow().date()
    today_total = round(sum(i.total for i in invoices if i.date.date() == today and not i.refunded), 2)
    today_count = sum(1 for i in invoices if i.date.date() == today)
    return templates.TemplateResponse(request, "invoices.html", {
        "staff": staff, "invoices": invoices, "today_total": today_total, "today_count": today_count,
    })


def get_shop_info(db: Session):
    return {
        "name": get_setting(db, "shop_name", "TechPro+"),
        "address": get_setting(db, "shop_address", ""),
        "phone": get_setting(db, "shop_phone", ""),
        "email": get_setting(db, "shop_email", ""),
        "gst": get_setting(db, "shop_gst", ""),
        "pst": get_setting(db, "shop_pst", ""),
    }


@app.get("/invoices/{invoice_id}", response_class=HTMLResponse)
def invoice_detail(request: Request, invoice_id: str, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    invoice = db.get(Invoice, invoice_id)
    if not invoice:
        return RedirectResponse("/invoices", status_code=303)
    tax_lines = json.loads(invoice.tax_breakdown) if invoice.tax_breakdown else []
    can_refund = role_allowed(staff, "owner", "manager")
    flash = request.session.pop("flash", None)
    province = get_setting(db, "province", "ON")
    all_products = db.query(Product).order_by(Product.name).all() if can_refund else []
    return templates.TemplateResponse(request, "invoice_detail.html", {
        "staff": staff, "invoice": invoice, "shop": get_shop_info(db),
        "province_label": PROVINCE_LABELS.get(province, province),
        "tax_lines": tax_lines, "can_refund": can_refund, "flash": flash,
        "all_products": all_products,
    })


@app.get("/invoices/{invoice_id}/thermal", response_class=HTMLResponse)
def invoice_thermal(request: Request, invoice_id: str, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    invoice = db.get(Invoice, invoice_id)
    if not invoice:
        return RedirectResponse("/invoices", status_code=303)
    tax_lines = json.loads(invoice.tax_breakdown) if invoice.tax_breakdown else []
    return templates.TemplateResponse(request, "invoice_thermal.html", {
        "staff": staff, "invoice": invoice, "shop": get_shop_info(db), "tax_lines": tax_lines,
    })


# ── LAYAWAY ──────────────────────────────────────────────────────────
@app.post("/pos/layaway/new")
def pos_layaway_new(request: Request, deposit: float = Form(0), due_date: str = Form(""),
                     db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    """Converts the current cart into a layaway instead of a straight
    sale. Stock is deducted immediately (reserved for this customer, so
    the same item can't be sold to someone else while it's being paid
    off) — the same way pos_checkout() deducts stock, just without
    creating an Invoice yet."""
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)

    cart = cart_get(request)
    if not cart:
        return RedirectResponse("/pos", status_code=303)

    customer_id = request.session.get("customer_id")
    customer = db.get(Customer, customer_id) if customer_id else None
    if not customer:
        request.session["scan_error"] = "A layaway needs a customer attached to the sale — select or add one first."
        return RedirectResponse("/pos", status_code=303)

    totals = cart_totals(request, db)
    deposit = round(max(0, deposit), 2)

    layaway = Layaway(
        number=next_layaway_number(db),
        customer_id=customer.id,
        cart_json=json.dumps(cart),
        subtotal=totals["sub"],
        tax_breakdown=json.dumps(totals["tax"]["lines"]),
        tax_total=totals["tax"]["tax_total"],
        total=totals["total"],
        paid_total=0,
        due_date=due_date.strip(),
        staff_id=staff.id,
    )
    db.add(layaway)
    db.flush()

    for item in cart:
        product = db.get(Product, item["product_id"]) if item.get("product_id") else None
        if product:
            product.stock = max(0, product.stock - item["qty"])

    if deposit > 0:
        db.add(LayawayPayment(layaway_id=layaway.id, amount=deposit, method="Cash", staff_id=staff.id, staff_name=staff.name))
        layaway.paid_total = deposit

    add_audit(db, staff, "LAYAWAY_CREATE", f"{layaway.number} — {customer.name} — total ${layaway.total:.2f}, deposit ${deposit:.2f}")
    db.commit()

    request.session["cart"] = []
    request.session["sub_override"] = None
    request.session["disc_value"] = 0
    request.session["disc_mode"] = "$"
    request.session["customer_id"] = None
    reset_customer_redemptions(request)
    return RedirectResponse(f"/layaway/{layaway.id}", status_code=303)


@app.get("/layaway", response_class=HTMLResponse)
def layaway_list(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    layaways = db.query(Layaway).order_by(Layaway.created_at.desc()).all()
    active_total_owed = round(sum(l.total - l.paid_total for l in layaways if l.status == "active"), 2)
    return templates.TemplateResponse(request, "layaway_list.html", {
        "staff": staff, "layaways": layaways, "active_total_owed": active_total_owed,
    })


@app.get("/layaway/{layaway_id}", response_class=HTMLResponse)
def layaway_detail(request: Request, layaway_id: str, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    layaway = db.get(Layaway, layaway_id)
    if not layaway:
        return RedirectResponse("/layaway", status_code=303)
    cart = json.loads(layaway.cart_json or "[]")
    tax_lines = json.loads(layaway.tax_breakdown or "[]")
    balance = round(layaway.total - layaway.paid_total, 2)
    return templates.TemplateResponse(request, "layaway_detail.html", {
        "staff": staff, "layaway": layaway, "cart": cart, "tax_lines": tax_lines,
        "balance": max(0, balance), "flash": request.session.pop("flash", None),
    })


@app.post("/layaway/{layaway_id}/payment")
def layaway_payment(request: Request, layaway_id: str, amount: float = Form(...), method: str = Form("Cash"),
                     db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    layaway = db.get(Layaway, layaway_id)
    if not layaway or layaway.status != "active":
        return RedirectResponse(f"/layaway/{layaway_id}", status_code=303)

    amount = round(amount, 2)
    if amount <= 0:
        request.session["flash"] = ("red", "Enter a payment amount greater than $0.")
        return RedirectResponse(f"/layaway/{layaway_id}", status_code=303)

    db.add(LayawayPayment(layaway_id=layaway.id, amount=amount, method=method, staff_id=staff.id, staff_name=staff.name))
    layaway.paid_total = round((layaway.paid_total or 0) + amount, 2)
    layaway.updated_at = datetime.utcnow()
    add_audit(db, staff, "LAYAWAY_PAYMENT", f"{layaway.number} — ${amount:.2f} via {method}")

    completed_note = ""
    if layaway.paid_total >= layaway.total - 0.005:
        # Fully paid off — convert to a real invoice. Stock was already
        # deducted at creation, so this does NOT touch product.stock again.
        cart = json.loads(layaway.cart_json or "[]")
        invoice = None
        for attempt in range(5):
            candidate = Invoice(
                number=next_invoice_number(db),
                customer_id=layaway.customer_id,
                staff_id=staff.id,
                payment_method=f"Layaway ({layaway.number})",
                # cash_amount/card_amount stay at their column default (0) here
                # on purpose: this one invoice represents the *whole* layaway,
                # but the actual cash/card money arrived across however many
                # separate LayawayPayment installments, each on its own day.
                # Cash Up counts those individually by their own created_at
                # date — attributing the full total to today (the payoff day)
                # would double-count every earlier installment and wreck the
                # reconciliation on both days.
                subtotal=layaway.subtotal,
                discount=0,
                tendered=layaway.paid_total,
                change_given=max(0, round(layaway.paid_total - layaway.total, 2)),
                tax_breakdown=layaway.tax_breakdown,
                tax_total=layaway.tax_total,
                total=layaway.total,
            )
            db.add(candidate)
            try:
                db.flush()
                invoice = candidate
                break
            except IntegrityError:
                db.rollback()
                if attempt == 4:
                    raise
        for item in cart:
            db.add(InvoiceLine(invoice_id=invoice.id, product_id=item.get("product_id"),
                                name=item["name"], sku=item.get("sku", ""), qty=item["qty"], price=item["price"],
                                imei=item.get("imei", "")))
        customer = db.get(Customer, layaway.customer_id) if layaway.customer_id else None
        if customer:
            points_per_dollar = float(get_setting(db, "points_per_dollar", "1"))
            customer.points = (customer.points or 0) + int(layaway.total * points_per_dollar)
            customer.spent = round((customer.spent or 0) + layaway.total, 2)
            customer.last_visit = datetime.utcnow().isoformat()
        layaway.status = "completed"
        layaway.invoice_id = invoice.id
        add_audit(db, staff, "LAYAWAY_COMPLETE", f"{layaway.number} paid off — converted to Invoice {invoice.number}")
        completed_note = f" Fully paid — converted to Invoice {invoice.number}."

    db.commit()
    request.session["flash"] = ("green", f"Payment of ${amount:.2f} recorded.{completed_note}")
    return RedirectResponse(f"/layaway/{layaway_id}", status_code=303)


@app.post("/layaway/{layaway_id}/cancel")
def layaway_cancel(request: Request, layaway_id: str, outcome: str = Form("cancelled"),
                    db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    """Ends an active layaway without completing the sale. Restocks the
    reserved items either way — 'cancelled' vs 'forfeited' only changes
    the record of what happened to the deposit already paid (forfeited =
    shop keeps it; cancelled = implies it gets refunded to the customer
    outside this system — refund the cash/store-credit manually)."""
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — cancelling a layaway requires manager or owner role.", status_code=403)

    layaway = db.get(Layaway, layaway_id)
    if layaway and layaway.status == "active":
        cart = json.loads(layaway.cart_json or "[]")
        for item in cart:
            product = db.get(Product, item.get("product_id")) if item.get("product_id") else None
            if product:
                product.stock = product.stock + item["qty"]
        layaway.status = "forfeited" if outcome == "forfeited" else "cancelled"
        layaway.updated_at = datetime.utcnow()
        add_audit(db, staff, "LAYAWAY_EDIT", f"{layaway.number} marked {layaway.status} — ${layaway.paid_total:.2f} paid so far, items restocked")
        db.commit()
        request.session["flash"] = ("green", f"Layaway {layaway.status}. Items have been restocked.")
    return RedirectResponse(f"/layaway/{layaway_id}", status_code=303)


@app.post("/invoices/{invoice_id}/refund")
def invoice_refund(request: Request, invoice_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — refunds require manager or owner role.", status_code=403)

    invoice = db.get(Invoice, invoice_id)
    if invoice and not invoice.refunded:
        invoice.refunded = True
        for line in invoice.lines:
            product = db.get(Product, line.product_id) if line.product_id else None
            if product:
                product.stock += line.qty
        if invoice.customer_id:
            customer = db.get(Customer, invoice.customer_id)
            if customer:
                customer.spent = round((customer.spent or 0) - invoice.total, 2)
        add_audit(db, staff, "REFUND", f"Refunded {invoice.number} — ${invoice.total:.2f}")
        db.commit()
    return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)


@app.post("/invoices/{invoice_id}/exchange-line/{line_id}")
def invoice_exchange_line(request: Request, invoice_id: str, line_id: str,
                           new_product_id: str = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    """Swaps one line item for a different product — the "exchange, not
    refund" flow shops with a no-refund policy actually need. Restocks
    the returned item, deducts the replacement from stock, and
    recalculates the invoice total, since price/tax may differ. Doesn't
    try to auto-process a second card/cash transaction for any price
    difference — that's simpler and more reliable handled at the
    register by the flash message telling staff the exact amount."""
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — exchanges require manager or owner role.", status_code=403)

    invoice = db.get(Invoice, invoice_id)
    if not invoice or invoice.refunded:
        return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)
    line = db.get(InvoiceLine, line_id)
    if not line or line.invoice_id != invoice_id:
        return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)
    new_product = db.get(Product, new_product_id)
    if not new_product:
        request.session["flash"] = ("red", "That replacement product couldn't be found.")
        return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)
    if new_product.stock < line.qty:
        request.session["flash"] = ("red", f"Only {new_product.stock} of {new_product.name} in stock — need {line.qty}.")
        return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)

    old_product = db.get(Product, line.product_id) if line.product_id else None
    if old_product:
        old_product.stock += line.qty  # returned item goes back on the shelf
    new_product.stock -= line.qty

    old_name, old_price = line.name, line.price
    line.exchange_note = (line.exchange_note + " | " if line.exchange_note else "") + \
        f"Exchanged from: {old_name} (${old_price:.2f}) on {datetime.utcnow().strftime('%b %d, %Y')}"
    line.product_id = new_product.id
    line.name = new_product.name
    line.sku = new_product.sku
    line.price = new_product.price

    # Recalculate the invoice total from scratch — price/tax may differ
    # between the old and new item.
    new_subtotal = round(sum(l.price * l.qty for l in invoice.lines), 2)
    taxable = max(0, new_subtotal - invoice.discount)
    province = get_setting(db, "province", "ON")
    tax = calc_canadian_tax(taxable, province)
    price_diff = round((new_product.price - old_price) * line.qty, 2)
    old_total = invoice.total
    invoice.subtotal = new_subtotal
    invoice.tax_breakdown = json.dumps(tax["lines"])
    invoice.tax_total = tax["tax_total"]
    invoice.total = round(taxable + tax["tax_total"], 2)

    add_audit(db, staff, "EXCHANGE",
              f"{invoice.number}: exchanged {old_name} for {new_product.name} (${price_diff:+.2f} difference)")
    db.commit()

    diff_vs_old_total = round(invoice.total - old_total, 2)
    if diff_vs_old_total > 0:
        msg = f"Exchange complete. Collect an additional ${diff_vs_old_total:.2f} from the customer."
    elif diff_vs_old_total < 0:
        msg = f"Exchange complete. Refund ${abs(diff_vs_old_total):.2f} to the customer (cash, card, or store credit)."
    else:
        msg = "Exchange complete. No price difference to collect or refund."
    request.session["flash"] = ("green", msg)
    return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)


@app.post("/invoices/{invoice_id}/email")
def invoice_email(request: Request, invoice_id: str, to_email: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    invoice = db.get(Invoice, invoice_id)
    if not invoice:
        return RedirectResponse("/invoices", status_code=303)
    recipient = to_email or (invoice.customer.email if invoice.customer else "")
    if not recipient:
        request.session["flash"] = ("red", "No email address on file for this customer.")
        return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)

    ok, msg = send_email_receipt(
        db, invoice, recipient, get_setting,
        province_label=PROVINCE_LABELS.get(get_setting(db, "province", "ON"), ""),
        base_url=str(request.base_url),
    )
    if ok:
        add_audit(db, staff, "EMAIL_RECEIPT", f"Receipt emailed to {recipient} for invoice {invoice.number}")
    request.session["flash"] = ("green" if ok else "red", msg)
    return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)


@app.post("/invoices/{invoice_id}/sms")
def invoice_sms(request: Request, invoice_id: str, to_phone: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    invoice = db.get(Invoice, invoice_id)
    if not invoice:
        return RedirectResponse("/invoices", status_code=303)
    recipient = to_phone or (invoice.customer.phone if invoice.customer else "")
    if not recipient:
        request.session["flash"] = ("red", "No phone number on file for this customer.")
        return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)

    shop_name = get_setting(db, "shop_name", "the shop")
    message = (f"Receipt from {shop_name} — Invoice {invoice.number}, total ${invoice.total:.2f}. "
               f"Thanks for your business!")
    ok, msg = send_sms(db, recipient, message, get_setting)
    if ok:
        add_audit(db, staff, "SMS_RECEIPT", f"Receipt SMS sent to {recipient} for invoice {invoice.number}")
        _log_outgoing_sms(db, recipient, message, staff.name, invoice.customer_id)
    request.session["flash"] = ("green" if ok else "red", msg)
    return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)


# ── PRODUCTS ─────────────────────────────────────────────────────────
def group_products_by_variant(products):
    """Groups products sharing a variant_group together (e.g. one 'Clear
    Silicone Case' entry covering many phone models/colors), leaving
    anything without a variant_group as standalone. Shared by the
    Products page and POS so both browse the same way. Tracks each
    group's full set of categories/brands across its variants, since a
    single case type can span multiple phone brands — the POS chip
    filters need to know a group matches if ANY variant inside it does,
    not just its first item."""
    groups = {}
    ungrouped = []
    for p in products:
        if p.variant_group:
            g = groups.setdefault(p.variant_group, {
                "name": p.variant_group, "variants": [], "total_stock": 0, "low_stock": False,
                "categories": set(), "brands": set(), "phone_brands": set(), "min_price": None, "max_price": None,
            })
            g["variants"].append(p)
            g["total_stock"] += p.stock
            if p.stock <= p.reorder_threshold:
                g["low_stock"] = True
            if p.category:
                g["categories"].add(p.category)
            if p.subcategory:
                g["brands"].add(p.subcategory)
            if p.phone_brand:
                g["phone_brands"].add(p.phone_brand)
            g["min_price"] = p.price if g["min_price"] is None else min(g["min_price"], p.price)
            g["max_price"] = p.price if g["max_price"] is None else max(g["max_price"], p.price)
        else:
            ungrouped.append(p)
    return sorted(groups.values(), key=lambda g: g["name"]), ungrouped


@app.get("/products", response_class=HTMLResponse)
def products_list(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    products = db.query(Product).order_by(Product.name).all()
    flash = request.session.pop("flash", None)
    last_generated_group = request.session.pop("last_generated_group", None)
    low_stock_count = sum(1 for p in products if p.stock <= p.reorder_threshold)
    stock_value = round(sum(p.cost * p.stock for p in products), 2) if role_allowed(staff, "owner", "manager") else None

    # Group variants (e.g. "iPhone 17 Cases" covering every case type/color
    # for that one model) together for browsing, without changing how
    # stock is tracked — each variant is still its own row with its own
    # SKU and stock underneath.
    product_groups, ungrouped = group_products_by_variant(products)

    # The actual browsing question staff have at the counter is "what do
    # we carry for THIS phone" — brand and model of the *phone*, not the
    # case's own maker. Built from phone_brand/phone_model, which only
    # phone-case products have set at all (chargers/cables/etc. won't
    # appear in this menu and stay reachable via search instead).
    phone_menu = {}
    for p in products:
        if p.phone_brand and p.phone_model:
            phone_menu.setdefault(p.phone_brand, set()).add(p.phone_model)
    phone_menu = {brand: sorted(models) for brand, models in phone_menu.items()}
    all_phone_brands = sorted(phone_menu.keys())

    return templates.TemplateResponse(request, "products.html", {
        "staff": staff, "products": products, "flash": flash,
        "last_generated_group": last_generated_group,
        "category_labels": CATEGORY_LABELS, "cat_subcategories": CAT_SUBCATEGORIES,
        "low_stock_count": low_stock_count, "stock_value": stock_value,
        "product_groups": product_groups, "ungrouped_products": ungrouped,
        "phone_menu": phone_menu, "all_phone_brands": all_phone_brands,
    })


@app.post("/products/import")
async def products_import(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — bulk import requires manager or owner role.", status_code=403)

    raw = (await file.read()).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))
    # Tolerant of different header casing/spacing, and matches our own CSV export columns
    def norm(d):
        return {k.strip().lower(): v for k, v in d.items()}

    existing_skus = {p.sku for p in db.query(Product).all() if p.sku}
    added, skipped, errors = 0, 0, 0
    for row in reader:
        row = norm(row)
        sku = (row.get("sku") or "").strip()
        name = (row.get("name") or "").strip()
        if not sku or not name:
            errors += 1
            continue
        if sku in existing_skus:
            skipped += 1
            continue
        try:
            price = float(row.get("price") or 0)
            cost = float(row.get("cost") or 0)
            stock = int(float(row.get("stock") or 0))
        except ValueError:
            errors += 1
            continue
        db.add(Product(sku=sku, name=name, category=(row.get("category") or "").strip(),
                        subcategory=(row.get("brand") or row.get("subcategory") or "").strip(),
                        variant_group=(row.get("variant group") or row.get("variant_group") or "").strip(),
                        price=price, cost=cost, stock=stock))
        existing_skus.add(sku)
        added += 1

    add_audit(db, staff, "PRODUCT_IMPORT", f"CSV import: {added} added, {skipped} skipped (duplicate SKU), {errors} invalid rows")
    db.commit()
    request.session["flash"] = ("green" if errors == 0 else "amber",
                                 f"Imported {added} new products. {skipped} skipped as duplicates. {errors} rows had errors.")
    return RedirectResponse("/products", status_code=303)


@app.post("/products/add")
def product_add(request: Request, sku: str = Form(...), name: str = Form(...),
                 category: str = Form(""), subcategory: str = Form(""), variant_group: str = Form(""),
                 price: float = Form(0), cost: float = Form(0), stock: int = Form(0),
                 reorder_threshold: int = Form(5), reorder_qty: int = Form(10), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — adding products requires manager or owner role.", status_code=403)
    db.add(Product(sku=sku, name=name, category=category, subcategory=subcategory, variant_group=variant_group.strip(),
                    price=price, cost=cost, stock=stock,
                    reorder_threshold=reorder_threshold, reorder_qty=reorder_qty))
    add_audit(db, staff, "PRODUCT_ADD", f"Added product: {name}")
    db.commit()
    return RedirectResponse("/products", status_code=303)


# ── BULK VARIANT GENERATOR ──────────────────────────────────────────
# The actual pain point this solves: a case style like "Hard Ring Case"
# needs its own SKU for every phone model x color combination it comes
# in — that's easily 15 models x 6 colors = 90 near-identical rows to
# create by hand through the one-at-a-time Add Product form. This lets
# staff pick the case style once, check off which models and colors it
# comes in, and generates every SKU in one pass.

_SKU_PART_RE = re.compile(r"[^A-Z0-9]+")


def _sku_part(text: str) -> str:
    """Slugifies free text into an uppercase SKU fragment — 'iPhone 15
    Pro Max' -> 'IPHONE15PROMAX', 'Rose Gold' -> 'ROSEGOLD'."""
    return _SKU_PART_RE.sub("", text.upper())


@app.get("/products/bulk-variants", response_class=HTMLResponse)
def bulk_variants_form(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — bulk-creating products requires manager or owner role.", status_code=403)
    return templates.TemplateResponse(request, "bulk_variants.html", {
        "staff": staff, "category_labels": CATEGORY_LABELS, "cat_subcategories": CAT_SUBCATEGORIES,
        "phone_brands": PHONE_BRANDS, "phone_models": PHONE_MODELS, "case_colors": CASE_COLORS,
        "common_case_styles": COMMON_CASE_STYLES,
        "laptop_models": LAPTOP_MODELS, "console_models": CONSOLE_MODELS,
        "common_laptop_gaming_styles": COMMON_LAPTOP_GAMING_STYLES,
        "flash": request.session.pop("flash", None),
    })


@app.post("/products/bulk-variants/generate")
def bulk_variants_generate(
    request: Request,
    variant_group: str = Form(...),
    category: str = Form("CASE"),
    subcategory: str = Form(""),
    sku_prefix: str = Form(""),
    price: float = Form(...),
    cost: float = Form(0),
    stock: int = Form(0),
    reorder_threshold: int = Form(5),
    reorder_qty: int = Form(10),
    phone_selections: list = Form([]),  # each entry: "Brand||Model", from checked checkboxes
    extra_models: str = Form(""),  # freeform textarea, one "Brand: Model" per line — for anything not in the checklist
    colors: list = Form([]),
    extra_colors: str = Form(""),  # freeform comma-separated — for shades not in the checklist
    db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect),
):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — bulk-creating products requires manager or owner role.", status_code=403)

    variant_group = variant_group.strip()

    phone_pairs = []
    for entry in phone_selections:
        if "||" in entry:
            brand, model = entry.split("||", 1)
            phone_pairs.append((brand.strip(), model.strip()))
    for line in extra_models.splitlines():
        line = line.strip()
        if not line:
            continue
        if ":" in line:
            brand, model = line.split(":", 1)
            phone_pairs.append((brand.strip(), model.strip()))
        else:
            phone_pairs.append(("Other", line))

    color_list = [c.strip() for c in colors if c.strip()]
    color_list += [c.strip() for c in extra_colors.split(",") if c.strip()]

    if not variant_group or not phone_pairs or not color_list:
        request.session["flash"] = ("red", "Pick a case style, at least one phone model, and at least one color before generating.")
        return RedirectResponse("/products/bulk-variants", status_code=303)

    prefix = _sku_part(sku_prefix or variant_group)[:14]
    existing_skus = {p.sku for p in db.query(Product).all() if p.sku}

    created, skipped = 0, 0
    for brand, model in phone_pairs:
        for color in color_list:
            sku = f"{prefix}-{_sku_part(model)[:16]}-{_sku_part(color)[:8]}"
            if sku in existing_skus:
                # Extremely similar names (e.g. "iPhone 15" and "iPhone
                # 15 Pro" slugging to overlapping fragments) can collide
                # — disambiguate rather than silently dropping a variant.
                suffix = 2
                while f"{sku}-{suffix}" in existing_skus:
                    suffix += 1
                sku = f"{sku}-{suffix}"
            name = f"{variant_group} — {model} ({color})"
            db.add(Product(
                sku=sku, name=name, category=category, subcategory=subcategory.strip(),
                variant_group=variant_group, phone_brand=brand, phone_model=model, color=color,
                price=price, cost=cost, stock=stock,
                reorder_threshold=reorder_threshold, reorder_qty=reorder_qty,
            ))
            existing_skus.add(sku)
            created += 1

    add_audit(db, staff, "PRODUCT_BULK_ADD",
              f"Bulk-generated {created} variants of \"{variant_group}\" across {len(phone_pairs)} models x {len(color_list)} colors")
    db.commit()

    request.session["flash"] = ("green", f'Created {created} new variants of "{variant_group}".')
    request.session["last_generated_group"] = variant_group
    return RedirectResponse("/products", status_code=303)


# ── BARCODE LABEL PRINTING ──────────────────────────────────────────
@app.get("/products/labels/print", response_class=HTMLResponse)
def product_labels_print(request: Request, db: Session = Depends(get_db),
                          group: str = "", ids: str = "", copies: int = 1):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)

    products = []
    if group:
        products = db.query(Product).filter(Product.variant_group == group).order_by(Product.name).all()
    elif ids:
        id_list = [i.strip() for i in ids.split(",") if i.strip()]
        # Preserve the order IDs were requested in, rather than whatever
        # order the DB happens to return them — matters when printing a
        # hand-picked set of labels one at a time.
        found = {p.id: p for p in db.query(Product).filter(Product.id.in_(id_list)).all()}
        products = [found[i] for i in id_list if i in found]

    copies = max(1, min(copies, 50))  # sane ceiling — a fat-fingered "500" shouldn't render 500 barcodes

    # Generate once per unique SKU, not once per copy — with copies=10
    # on a 20-variant batch that's the difference between 20 and 200
    # barcode generations for the exact same visual output.
    barcode_by_sku = {}
    for p in products:
        if p.sku not in barcode_by_sku:
            svg, width_mm = generate_barcode_svg(p.sku)
            barcode_by_sku[p.sku] = {"svg": svg, "width_mm": width_mm}

    return templates.TemplateResponse(request, "product_labels.html", {
        "staff": staff, "products": products, "group": group, "ids": ids, "copies": copies,
        "shop": get_shop_info(db), "barcode_by_sku": barcode_by_sku,
    })


@app.get("/products/reorder", response_class=HTMLResponse)
def reorder_list(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — the reorder list requires manager or owner role.", status_code=403)
    low_stock = (db.query(Product)
                 .filter(Product.stock <= Product.reorder_threshold)
                 .order_by(Product.stock).all())
    total_cost = round(sum(p.cost * p.reorder_qty for p in low_stock), 2)
    return templates.TemplateResponse(request, "reorder_list.html", {
        "staff": staff, "products": low_stock, "total_cost": total_cost,
    })


@app.get("/products/reorder/print", response_class=HTMLResponse)
def reorder_list_print(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — the reorder list requires manager or owner role.", status_code=403)
    low_stock = (db.query(Product)
                 .filter(Product.stock <= Product.reorder_threshold)
                 .order_by(Product.category, Product.name).all())
    total_cost = round(sum(p.cost * p.reorder_qty for p in low_stock), 2)
    shop_name = get_setting(db, "shop_name", "TechPro+")
    return templates.TemplateResponse(request, "reorder_print.html", {
        "staff": staff, "products": low_stock, "total_cost": total_cost, "shop_name": shop_name,
    })


@app.get("/export/csv/reorder")
def export_reorder_csv(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — exports require manager or owner role.", status_code=403)
    low_stock = db.query(Product).filter(Product.stock <= Product.reorder_threshold).order_by(Product.name).all()
    rows = [(p.sku, p.name, p.category, p.subcategory, p.stock, p.reorder_threshold, p.reorder_qty, round(p.cost * p.reorder_qty, 2)) for p in low_stock]
    return _csv_response(rows, ["SKU", "Name", "Category", "Brand", "Current Stock", "Reorder At", "Suggested Qty", "Est. Cost"], "reorder_list.csv")


# ── SUPPLIERS ────────────────────────────────────────────────────────────
@app.get("/suppliers", response_class=HTMLResponse)
def suppliers_list(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — suppliers require manager or owner role.", status_code=403)
    suppliers = db.query(Supplier).order_by(Supplier.name).all()
    return templates.TemplateResponse(request, "suppliers.html", {"staff": staff, "suppliers": suppliers})


@app.post("/suppliers/add")
def supplier_add(request: Request, name: str = Form(...), contact_name: str = Form(""),
                  email: str = Form(""), phone: str = Form(""), lead_time_days: int = Form(7),
                  notes: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    db.add(Supplier(name=name, contact_name=contact_name, email=email, phone=phone,
                     lead_time_days=lead_time_days, notes=notes, active=True))
    add_audit(db, staff, "SUPPLIER_ADD", f"Added supplier: {name}")
    db.commit()
    return RedirectResponse("/suppliers", status_code=303)


@app.post("/suppliers/{supplier_id}/edit")
def supplier_edit(request: Request, supplier_id: str, name: str = Form(...), contact_name: str = Form(""),
                   email: str = Form(""), phone: str = Form(""), lead_time_days: int = Form(7),
                   notes: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    supplier = db.get(Supplier, supplier_id)
    if supplier:
        supplier.name, supplier.contact_name = name, contact_name
        supplier.email, supplier.phone = email, phone
        supplier.lead_time_days, supplier.notes = lead_time_days, notes
        add_audit(db, staff, "SUPPLIER_EDIT", f"Edited supplier: {name}")
        db.commit()
    return RedirectResponse("/suppliers", status_code=303)


@app.post("/suppliers/{supplier_id}/toggle")
def supplier_toggle(request: Request, supplier_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    supplier = db.get(Supplier, supplier_id)
    if supplier:
        supplier.active = not supplier.active
        add_audit(db, staff, "SUPPLIER_TOGGLE", f"{'Activated' if supplier.active else 'Deactivated'} supplier: {supplier.name}")
        db.commit()
    return RedirectResponse("/suppliers", status_code=303)


# ── PURCHASE ORDERS ──────────────────────────────────────────────────────
def _next_po_number(db: Session) -> str:
    count = db.query(PurchaseOrder).count()
    return f"PO-{count + 1:04d}"


@app.get("/purchase-orders", response_class=HTMLResponse)
def po_list(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — purchase orders require manager or owner role.", status_code=403)
    pos = db.query(PurchaseOrder).order_by(PurchaseOrder.created_at.desc()).all()
    po_totals = {po.id: round(sum(l.qty * l.unit_cost for l in po.lines), 2) for po in pos}
    return templates.TemplateResponse(request, "purchase_orders.html", {"staff": staff, "pos": pos, "po_totals": po_totals})


@app.get("/purchase-orders/new", response_class=HTMLResponse)
def po_new(request: Request, from_reorder: str = "", db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — purchase orders require manager or owner role.", status_code=403)
    suppliers = db.query(Supplier).filter(Supplier.active == True).order_by(Supplier.name).all()  # noqa: E712
    prefill = []
    if from_reorder:
        low_stock = db.query(Product).filter(Product.stock <= Product.reorder_threshold).order_by(Product.name).all()
        prefill = [{"id": p.id, "sku": p.sku, "name": p.name, "qty": p.reorder_qty, "cost": p.cost} for p in low_stock]
    all_products = db.query(Product).order_by(Product.name).all()
    return templates.TemplateResponse(request, "po_new.html", {
        "staff": staff, "suppliers": suppliers, "prefill": prefill, "all_products": all_products,
    })


@app.post("/purchase-orders/add")
async def po_add(request: Request, supplier_id: str = Form(""), notes: str = Form(""),
                  db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    form = await request.form()
    product_ids = form.getlist("product_id")
    qtys = form.getlist("qty")
    costs = form.getlist("unit_cost")

    po = PurchaseOrder(number=_next_po_number(db), supplier_id=supplier_id or None,
                        status="draft", created_by_name=staff.name, notes=notes)
    db.add(po)
    db.flush()

    line_count = 0
    for pid, qty_raw, cost_raw in zip(product_ids, qtys, costs):
        if not pid or not qty_raw:
            continue
        try:
            qty, cost = int(qty_raw), float(cost_raw or 0)
        except ValueError:
            continue
        if qty <= 0:
            continue
        product = db.get(Product, pid)
        if not product:
            continue
        db.add(PurchaseOrderLine(po_id=po.id, product_id=product.id, name=product.name,
                                  sku=product.sku, qty=qty, unit_cost=cost))
        line_count += 1

    if line_count == 0:
        db.rollback()
        request.session["flash"] = ("red", "No valid line items — purchase order was not created.")
        return RedirectResponse("/purchase-orders/new", status_code=303)

    add_audit(db, staff, "PO_CREATE", f"Created {po.number} with {line_count} line item(s)")
    db.commit()
    return RedirectResponse(f"/purchase-orders/{po.id}", status_code=303)


@app.get("/purchase-orders/{po_id}", response_class=HTMLResponse)
def po_detail(request: Request, po_id: str, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    po = db.get(PurchaseOrder, po_id)
    if not po:
        return RedirectResponse("/purchase-orders", status_code=303)
    total_cost = round(sum(l.qty * l.unit_cost for l in po.lines), 2)
    flash = request.session.pop("flash", None)
    return templates.TemplateResponse(request, "po_detail.html", {
        "staff": staff, "po": po, "total_cost": total_cost, "flash": flash,
    })


@app.post("/purchase-orders/{po_id}/send")
def po_send(request: Request, po_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    po = db.get(PurchaseOrder, po_id)
    if po and po.status == "draft":
        po.status = "sent"
        po.sent_at = datetime.utcnow()
        add_audit(db, staff, "PO_SEND", f"Marked {po.number} as sent")

        email_note = ""
        if po.supplier and po.supplier.email:
            lines_text = "\n".join(f"{l.sku} — {l.name} x{l.qty} @ ${l.unit_cost:.2f}" for l in po.lines)
            body = (f"Purchase Order {po.number}\n\n{lines_text}\n\n"
                    f"Total: ${sum(l.qty * l.unit_cost for l in po.lines):.2f}\n\n{po.notes}")
            ok, detail = send_plain_email(db, po.supplier.email, f"Purchase Order {po.number}", body, get_setting)
            email_note = f" Email to supplier: {detail}" if ok else f" Email failed: {detail}"
        db.commit()
        request.session["flash"] = ("green", f"{po.number} marked as sent.{email_note}")
    return RedirectResponse(f"/purchase-orders/{po_id}", status_code=303)


@app.post("/purchase-orders/{po_id}/receive")
async def po_receive(request: Request, po_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    po = db.get(PurchaseOrder, po_id)
    if not po or po.status == "received":
        return RedirectResponse(f"/purchase-orders/{po_id}", status_code=303)

    form = await request.form()
    received_in = 0
    for line in po.lines:
        field = f"received_{line.id}"
        raw = form.get(field, "")
        try:
            received_qty = int(raw) if raw else line.qty  # default: assume full qty received if left blank
        except ValueError:
            received_qty = 0
        received_qty = max(0, min(received_qty, line.qty))
        line.received_qty = received_qty
        if line.product_id and received_qty > 0:
            product = db.get(Product, line.product_id)
            if product:
                product.stock += received_qty
                received_in += received_qty

    po.status = "received"
    po.received_at = datetime.utcnow()
    add_audit(db, staff, "PO_RECEIVE", f"Received {po.number} — {received_in} unit(s) added to stock")
    db.commit()
    request.session["flash"] = ("green", f"Stock updated — {received_in} unit(s) received into inventory.")
    return RedirectResponse(f"/purchase-orders/{po_id}", status_code=303)


@app.get("/products/{product_id}/edit", response_class=HTMLResponse)
def product_edit_page(request: Request, product_id: str, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — editing products requires manager or owner role.", status_code=403)
    product = db.get(Product, product_id)
    if not product:
        return RedirectResponse("/products", status_code=303)
    return templates.TemplateResponse(request, "product_edit.html", {
        "staff": staff, "product": product,
        "category_labels": CATEGORY_LABELS, "cat_subcategories": CAT_SUBCATEGORIES,
    })


@app.post("/products/{product_id}/edit")
def product_edit(request: Request, product_id: str, name: str = Form(...), sku: str = Form(...),
                  category: str = Form(""), subcategory: str = Form(""), variant_group: str = Form(""),
                  price: float = Form(0), cost: float = Form(0), stock: int = Form(0),
                  reorder_threshold: int = Form(5), reorder_qty: int = Form(10), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — editing products requires manager or owner role.", status_code=403)
    product = db.get(Product, product_id)
    if product:
        product.name, product.sku = name, sku
        product.category, product.subcategory, product.variant_group = category, subcategory, variant_group.strip()
        product.price, product.cost, product.stock = price, cost, stock
        product.reorder_threshold, product.reorder_qty = reorder_threshold, reorder_qty
        add_audit(db, staff, "PRODUCT_EDIT", f"Edited product: {name}")
        db.commit()
    return RedirectResponse("/products", status_code=303)


@app.post("/products/{product_id}/delete")
def product_delete(request: Request, product_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    product = db.get(Product, product_id)
    if product:
        add_audit(db, staff, "PRODUCT_DELETE", f"Deleted product: {product.name}")
        db.delete(product)
        db.commit()
    return RedirectResponse("/products", status_code=303)


# ── CUSTOMERS ────────────────────────────────────────────────────────
@app.get("/customers", response_class=HTMLResponse)
def customers_list(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    customers = db.query(Customer).order_by(Customer.name).all()
    total_store_credit = round(sum(c.store_credit for c in customers), 2)
    total_points = sum(c.points for c in customers)
    return templates.TemplateResponse(request, "customers.html", {
        "staff": staff, "customers": customers,
        "total_store_credit": total_store_credit, "total_points": total_points,
    })


@app.post("/customers/add")
def customer_add(request: Request, name: str = Form(...), phone: str = Form(""),
                  email: str = Form(""), notes: str = Form(""),
                  marketing_consent: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    consented = marketing_consent == "on"
    db.add(Customer(name=name, phone=phone, email=email, notes=notes,
                     marketing_consent=consented, consent_date=datetime.utcnow() if consented else None))
    add_audit(db, staff, "CUSTOMER_ADD", f"Added customer: {name}")
    db.commit()
    return RedirectResponse("/customers", status_code=303)


@app.get("/customers/{customer_id}", response_class=HTMLResponse)
def customer_detail(request: Request, customer_id: str, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    customer = db.get(Customer, customer_id)
    if not customer:
        return RedirectResponse("/customers", status_code=303)
    invoices = db.query(Invoice).filter(Invoice.customer_id == customer_id).order_by(Invoice.date.desc()).all()
    repairs = db.query(Repair).filter(Repair.customer_id == customer_id).order_by(Repair.created_at.desc()).all()
    repair_warranties = {r.id: get_warranty_status(r) for r in repairs}
    sms_thread = db.query(SmsMessage).filter(SmsMessage.customer_id == customer_id).order_by(SmsMessage.created_at).all()
    trade_ins = db.query(TradeIn).filter(TradeIn.customer_id == customer_id).order_by(TradeIn.created_at.desc()).all()
    layaways = db.query(Layaway).filter(Layaway.customer_id == customer_id).order_by(Layaway.created_at.desc()).all()
    flash = request.session.pop("flash", None)
    return templates.TemplateResponse(request, "customer_detail.html", {
        "staff": staff, "customer": customer, "invoices": invoices,
        "repairs": repairs, "repair_warranties": repair_warranties,
        "status_labels": STATUS_LABELS, "status_badge": STATUS_BADGE,
        "sms_thread": sms_thread, "flash": flash,
        "trade_ins": trade_ins, "layaways": layaways,
    })


@app.post("/customers/{customer_id}/sms")
def customer_send_sms(request: Request, customer_id: str, message: str = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    customer = db.get(Customer, customer_id)
    if not customer or not customer.phone:
        request.session["flash"] = ("red", "No phone number on file for this customer.")
        return RedirectResponse(f"/customers/{customer_id}", status_code=303)

    ok, msg = send_sms(db, customer.phone, message, get_setting)
    if ok:
        _log_outgoing_sms(db, customer.phone, message, staff.name, customer.id)
        add_audit(db, staff, "SMS_SENT", f"SMS sent to {customer.name}")
    request.session["flash"] = ("green" if ok else "red", msg)
    return RedirectResponse(f"/customers/{customer_id}", status_code=303)


@app.post("/customers/{customer_id}/edit")
def customer_edit(request: Request, customer_id: str, name: str = Form(...),
                   phone: str = Form(""), email: str = Form(""),
                   marketing_consent: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    customer = db.get(Customer, customer_id)
    if customer:
        customer.name, customer.phone, customer.email = name, phone, email
        newly_consented = marketing_consent == "on"
        if newly_consented and not customer.marketing_consent:
            customer.consent_date = datetime.utcnow()  # only stamp the date when consent is newly given
        customer.marketing_consent = newly_consented
        add_audit(db, staff, "CUSTOMER_EDIT", f"Edited customer: {name}")
        db.commit()
    return RedirectResponse(f"/customers/{customer_id}", status_code=303)


@app.post("/customers/{customer_id}/notes")
def customer_notes(request: Request, customer_id: str, notes: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    customer = db.get(Customer, customer_id)
    if customer:
        customer.notes = notes
        add_audit(db, staff, "CUSTOMER_EDIT", f"Updated notes for {customer.name}")
        db.commit()
    return RedirectResponse(f"/customers/{customer_id}", status_code=303)


@app.post("/customers/{customer_id}/credit")
def customer_credit(request: Request, customer_id: str, amount: float = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — issuing store credit requires manager or owner role.", status_code=403)
    customer = db.get(Customer, customer_id)
    if customer:
        customer.store_credit = round((customer.store_credit or 0) + amount, 2)
        add_audit(db, staff, "STORE_CREDIT", f"Issued ${amount:.2f} store credit to {customer.name}")
        db.commit()
    return RedirectResponse(f"/customers/{customer_id}", status_code=303)


# ── TRADE-INS ────────────────────────────────────────────────────────
@app.get("/trade-ins", response_class=HTMLResponse)
def trade_ins_list(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    trade_ins = db.query(TradeIn).order_by(TradeIn.created_at.desc()).all()
    total_paid_out = round(sum(t.offered_amount for t in trade_ins), 2)
    customers = db.query(Customer).order_by(Customer.name).all()
    return templates.TemplateResponse(request, "trade_ins.html", {
        "staff": staff, "trade_ins": trade_ins, "total_paid_out": total_paid_out,
        "customers": customers, "flash": request.session.pop("flash", None),
    })


@app.post("/trade-ins/add")
def trade_in_add(request: Request, customer_id: str = Form(""), new_customer_name: str = Form(""),
                  new_customer_phone: str = Form(""), device: str = Form(...), imei: str = Form(""),
                  condition: str = Form(""), offered_amount: float = Form(...),
                  payout_method: str = Form("store_credit"), notes: str = Form(""),
                  db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager", "cashier"):
        return HTMLResponse("Forbidden.", status_code=403)

    customer = db.get(Customer, customer_id) if customer_id else None
    if not customer and new_customer_name.strip():
        customer = Customer(name=new_customer_name.strip(), phone=new_customer_phone.strip())
        db.add(customer)
        db.flush()

    trade_in = TradeIn(
        customer_id=customer.id if customer else None,
        device=device.strip(), imei=imei.strip(), condition=condition.strip(),
        offered_amount=round(offered_amount, 2), payout_method=payout_method,
        notes=notes.strip(), staff_id=staff.id, staff_name=staff.name,
    )
    db.add(trade_in)

    if customer and payout_method == "store_credit" and offered_amount > 0:
        customer.store_credit = round((customer.store_credit or 0) + offered_amount, 2)
        add_audit(db, staff, "STORE_CREDIT", f"Issued ${offered_amount:.2f} store credit for trade-in ({device}) — {customer.name}")

    add_audit(db, staff, "TRADE_IN", f"Accepted trade-in: {device}" + (f" (IMEI {imei.strip()})" if imei.strip() else "") + f" — ${offered_amount:.2f} via {payout_method.replace('_', ' ')}")
    db.commit()

    if payout_method == "cash":
        request.session["flash"] = ("green", f"Trade-in recorded. Pay ${offered_amount:.2f} cash out from the till.")
    else:
        request.session["flash"] = ("green", f"Trade-in recorded — ${offered_amount:.2f} store credit issued{' to ' + customer.name if customer else ''}.")
    return RedirectResponse("/trade-ins", status_code=303)


@app.post("/trade-ins/{trade_in_id}/status")
def trade_in_status(request: Request, trade_in_id: str, status: str = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    trade_in = db.get(TradeIn, trade_in_id)
    if trade_in and status in ("accepted", "resold", "scrapped"):
        trade_in.status = status
        add_audit(db, staff, "TRADE_IN_EDIT", f"{trade_in.device} marked {status}")
        db.commit()
    return RedirectResponse("/trade-ins", status_code=303)


# ── REPAIRS ──────────────────────────────────────────────────────────
@app.get("/repairs", response_class=HTMLResponse)
def repairs_list(request: Request, view: str = "kanban", db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    repairs = db.query(Repair).order_by(Repair.created_at.desc()).all()
    columns = {s: [] for s in STATUS_ORDER}
    for r in repairs:
        columns.setdefault(r.status, []).append(r)
    technicians = db.query(Staff).filter(Staff.active == True).all()  # noqa: E712
    customers = db.query(Customer).order_by(Customer.name).all()
    return templates.TemplateResponse(request, "repairs.html", {
        "staff": staff, "repairs": repairs, "columns": columns, "view": view,
        "status_labels": STATUS_LABELS, "status_order": STATUS_ORDER, "status_badge": STATUS_BADGE,
        "issue_types": ISSUE_TYPES, "technicians": technicians, "customers": customers,
        "today": datetime.utcnow().date().isoformat(),
    })


@app.post("/repairs/add")
def repair_add(request: Request, phone: str = Form(...), name: str = Form(...),
                device: str = Form(...), imei: str = Form(""), issue: str = Form(...), description: str = Form(""),
                estimated_cost: str = Form(""), warranty_days: int = Form(90),
                promised_by: str = Form(""), technician_id: str = Form(""),
                db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)

    customer = db.query(Customer).filter(Customer.phone == phone).first()
    if not customer:
        customer = Customer(name=name, phone=phone)
        db.add(customer)
        db.flush()

    last = db.query(Repair).order_by(Repair.ticket_no.desc()).first()
    next_ticket = (last.ticket_no + 1) if last else 1001

    cost_val = float(estimated_cost) if estimated_cost else None
    history = [{"status": "RECEIVED", "note": "Ticket created", "date": datetime.utcnow().isoformat()}]

    repair = Repair(
        ticket_no=next_ticket, customer_id=customer.id, device=device, imei=imei.strip(), issue=issue,
        description=description, status="RECEIVED", estimated_cost=cost_val,
        warranty_days=warranty_days, promised_by=promised_by,
        technician_id=technician_id or None, status_history=json.dumps(history),
    )
    db.add(repair)
    add_audit(db, staff, "REPAIR_CREATE", f"Ticket #{next_ticket} — {device} ({issue})" + (f" — IMEI {imei.strip()}" if imei.strip() else ""))
    db.commit()
    return RedirectResponse("/repairs", status_code=303)


@app.post("/repairs/{repair_id}/imei")
def repair_imei(request: Request, repair_id: str, imei: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    repair = db.get(Repair, repair_id)
    if repair:
        repair.imei = imei.strip()
        add_audit(db, staff, "REPAIR_EDIT", f"#{repair.ticket_no} — IMEI/serial set to {imei.strip() or '(cleared)'}")
        db.commit()
    return RedirectResponse(f"/repairs/{repair_id}", status_code=303)


@app.get("/repairs/{repair_id}", response_class=HTMLResponse)
def repair_detail(request: Request, repair_id: str, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    repair = db.get(Repair, repair_id)
    if not repair:
        return RedirectResponse("/repairs", status_code=303)
    history = json.loads(repair.status_history) if repair.status_history else []
    n_status = next_status(repair.status)
    cur_idx = STATUS_ORDER.index(repair.status) if repair.status in STATUS_ORDER else 0
    technicians = db.query(Staff).filter(Staff.active == True).all()  # noqa: E712
    flash = request.session.pop("flash", None)
    linked_invoices = db.query(Invoice).filter(Invoice.repair_id == repair_id).order_by(Invoice.date.desc()).all()

    parts = db.query(RepairPart).filter(RepairPart.repair_id == repair_id).order_by(RepairPart.added_at).all()
    parts_cost = round(sum(p.qty * p.unit_cost for p in parts), 2)
    margin = None
    if role_allowed(staff, "owner", "manager"):
        charge = repair.final_cost if repair.final_cost is not None else (repair.estimated_cost or 0)
        margin = round(charge - parts_cost, 2)
    all_products = db.query(Product).order_by(Product.name).all() if role_allowed(staff, "owner", "manager", "technician") else []
    warranty = get_warranty_status(repair)

    return templates.TemplateResponse(request, "repair_detail.html", {
        "staff": staff, "repair": repair, "history": list(reversed(history)),
        "next_status": n_status, "cur_idx": cur_idx,
        "status_labels": STATUS_LABELS, "status_order": STATUS_ORDER, "status_badge": STATUS_BADGE,
        "technicians": technicians, "flash": flash, "linked_invoices": linked_invoices, "warranty": warranty,
        "parts": parts, "parts_cost": parts_cost, "margin": margin, "all_products": all_products,
    })


@app.post("/repairs/{repair_id}/parts/add")
def repair_part_add(request: Request, repair_id: str, product_id: str = Form(...),
                     qty: int = Form(1), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager", "technician"):
        return HTMLResponse("Forbidden — adding parts requires technician, manager, or owner role.", status_code=403)
    repair = db.get(Repair, repair_id)
    product = db.get(Product, product_id)
    if not repair or not product:
        return RedirectResponse(f"/repairs/{repair_id}", status_code=303)
    if qty <= 0:
        request.session["flash"] = ("red", "Quantity must be at least 1.")
        return RedirectResponse(f"/repairs/{repair_id}", status_code=303)
    if product.stock < qty:
        request.session["flash"] = ("red", f"Only {product.stock} of {product.name} in stock — can't use {qty}.")
        return RedirectResponse(f"/repairs/{repair_id}", status_code=303)

    db.add(RepairPart(repair_id=repair_id, product_id=product.id, name=product.name,
                       qty=qty, unit_cost=product.cost))
    product.stock -= qty  # deducted immediately, same as a POS sale — parts don't silently vanish from counts
    add_audit(db, staff, "REPAIR_PART_ADD", f"Used {qty}x {product.name} on ticket #{repair.ticket_no}")
    db.commit()
    request.session["flash"] = ("green", f"Added {qty}x {product.name} — stock updated.")
    return RedirectResponse(f"/repairs/{repair_id}", status_code=303)


@app.post("/repairs/{repair_id}/parts/{part_id}/remove")
def repair_part_remove(request: Request, repair_id: str, part_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager", "technician"):
        return HTMLResponse("Forbidden", status_code=403)
    part = db.get(RepairPart, part_id)
    if part and part.repair_id == repair_id:
        if part.product_id:
            product = db.get(Product, part.product_id)
            if product:
                product.stock += part.qty  # restore stock — this part wasn't actually used after all
        add_audit(db, staff, "REPAIR_PART_REMOVE", f"Removed {part.qty}x {part.name} from ticket, stock restored")
        db.delete(part)
        db.commit()
        request.session["flash"] = ("green", "Part removed and stock restored.")
    return RedirectResponse(f"/repairs/{repair_id}", status_code=303)


@app.post("/repairs/{repair_id}/advance")
def repair_advance(request: Request, repair_id: str, note: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    repair = db.get(Repair, repair_id)
    if repair:
        n = next_status(repair.status)
        if n:
            old = repair.status
            repair.status = n
            repair.updated_at = datetime.utcnow()
            history = json.loads(repair.status_history) if repair.status_history else []
            history.append({"status": n, "note": note or f"Moved from {STATUS_LABELS.get(old, old)}", "date": datetime.utcnow().isoformat()})
            repair.status_history = json.dumps(history)
            add_audit(db, staff, "REPAIR_STATUS", f"#{repair.ticket_no} → {STATUS_LABELS.get(n, n)}")
            db.commit()
    return RedirectResponse(f"/repairs/{repair_id}", status_code=303)


@app.post("/repairs/{repair_id}/notify")
def repair_notify(request: Request, repair_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    repair = db.get(Repair, repair_id)
    if not repair or not repair.customer or not repair.customer.phone:
        request.session["flash"] = ("red", "No customer phone number on file for this ticket.")
        return RedirectResponse(f"/repairs/{repair_id}", status_code=303)

    shop_name = get_setting(db, "shop_name", "the shop")
    message = (f"Hi {repair.customer.name}, great news! Your {repair.device} repair is "
               f"complete and ready for pickup at {shop_name}. Ticket #{repair.ticket_no}. See you soon!")
    ok, msg = send_sms(db, repair.customer.phone, message, get_setting)
    if ok:
        add_audit(db, staff, "REPAIR_NOTIFY", f"Ready-for-pickup SMS sent for ticket #{repair.ticket_no}")
        _log_outgoing_sms(db, repair.customer.phone, message, staff.name, repair.customer_id)
    request.session["flash"] = ("green" if ok else "red", msg)
    return RedirectResponse(f"/repairs/{repair_id}", status_code=303)


@app.post("/repairs/{repair_id}/cost")
def repair_cost(request: Request, repair_id: str, estimated_cost: str = Form(""),
                 final_cost: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    repair = db.get(Repair, repair_id)
    if repair:
        repair.estimated_cost = float(estimated_cost) if estimated_cost else repair.estimated_cost
        repair.final_cost = float(final_cost) if final_cost else repair.final_cost
        repair.updated_at = datetime.utcnow()
        add_audit(db, staff, "REPAIR_EDIT", f"Updated costs for ticket #{repair.ticket_no}")
        db.commit()
    return RedirectResponse(f"/repairs/{repair_id}", status_code=303)


@app.post("/repairs/{repair_id}/charge")
def repair_charge(request: Request, repair_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    """Send a repair ticket to the POS register to collect payment. Clears
    whatever's currently in the cart, adds a single line item for this
    repair (using the final cost if set, otherwise the estimate), attaches
    the customer, and remembers the repair via session so /pos/checkout
    can link the resulting invoice back to this ticket."""
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    repair = db.get(Repair, repair_id)
    if not repair:
        return RedirectResponse("/repairs", status_code=303)

    price = repair.final_cost if repair.final_cost is not None else (repair.estimated_cost or 0)
    request.session["cart"] = [{
        "product_id": "",
        "name": f"Repair #{repair.ticket_no} — {repair.device} ({repair.issue})",
        "sku": f"RPR-{repair.ticket_no}",
        "price": round(price, 2),
        "qty": 1,
    }]
    request.session["sub_override"] = None
    request.session["disc_value"] = 0
    request.session["disc_mode"] = "$"
    request.session["customer_id"] = repair.customer_id
    reset_customer_redemptions(request)
    request.session["pos_repair_id"] = repair.id
    return RedirectResponse("/pos", status_code=303)


# ── STAFF ────────────────────────────────────────────────────────────
@app.get("/staff", response_class=HTMLResponse)
def staff_list(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — staff management requires manager or owner role.", status_code=403)
    all_staff = db.query(Staff).all()
    security_question = get_setting(db, "security_question", "")
    security_configured = bool(get_setting(db, "security_answer_hash", ""))
    flash = request.session.pop("flash", None)
    return templates.TemplateResponse(request, "staff.html", {
        "staff": staff, "all_staff": all_staff, "flash": flash,
        "security_question": security_question, "security_configured": security_configured,
    })


@app.post("/staff/add")
def staff_add(request: Request, name: str = Form(...), pin: str = Form(...),
              role: str = Form("cashier"), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    current = require_login(request, db)
    if not current:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(current, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    if role == "owner" and current.role != "owner":
        add_audit(db, current, "SECURITY_BLOCK", f"Manager attempted to create an Owner account: {name}")
        db.commit()
        return HTMLResponse("Forbidden — only an Owner can create another Owner account.", status_code=403)
    if not pin.isdigit() or len(pin) != 4:
        return HTMLResponse("PIN must be exactly 4 digits", status_code=400)
    db.add(Staff(name=name, role=role, pin_hash=hash_pin(pin), active=True))
    add_audit(db, current, "STAFF_ADD", f"Added staff member: {name}")
    db.commit()
    return RedirectResponse("/staff", status_code=303)


@app.post("/staff/{staff_id}/edit")
def staff_edit(request: Request, staff_id: str, name: str = Form(...),
                role: str = Form(...), new_pin: str = Form(""),
                security_answer: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    current = require_login(request, db)
    if not current:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(current, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    target = db.get(Staff, staff_id)
    if not target:
        return RedirectResponse("/staff", status_code=303)

    if current.role != "owner" and target.role == "owner":
        add_audit(db, current, "SECURITY_BLOCK", f"Manager attempted to edit Owner account: {target.name}")
        db.commit()
        return HTMLResponse("Forbidden — only an Owner can edit an Owner account.", status_code=403)
    if current.role != "owner" and role == "owner":
        add_audit(db, current, "SECURITY_BLOCK", f"Manager attempted to grant Owner role to: {name}")
        db.commit()
        return HTMLResponse("Forbidden — only an Owner can grant the Owner role.", status_code=403)

    pin_changing = bool(new_pin and new_pin.isdigit() and len(new_pin) == 4)
    answer_hash = get_setting(db, "security_answer_hash", "")
    if pin_changing and answer_hash:
        if not verify_pin(security_answer.strip().lower(), answer_hash):
            add_audit(db, current, "SECURITY_BLOCK", f"Wrong security answer while changing PIN for: {target.name}")
            db.commit()
            request.session["flash"] = ("red", "Incorrect security answer — PIN was not changed.")
            return RedirectResponse("/staff", status_code=303)

    target.name, target.role = name, role
    if pin_changing:
        target.pin_hash = hash_pin(new_pin)
        add_audit(db, current, "PIN_CHANGE", f"Changed PIN for: {target.name}")
    add_audit(db, current, "STAFF_EDIT", f"Edited staff member: {name}")
    db.commit()
    return RedirectResponse("/staff", status_code=303)


@app.post("/staff/{staff_id}/toggle")
def staff_toggle(request: Request, staff_id: str, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    current = require_login(request, db)
    if not current:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(current, "owner", "manager"):
        return HTMLResponse("Forbidden", status_code=403)
    target = db.get(Staff, staff_id)
    if not target:
        return RedirectResponse("/staff", status_code=303)

    if current.role != "owner" and target.role == "owner":
        add_audit(db, current, "SECURITY_BLOCK", f"Manager attempted to deactivate Owner account: {target.name}")
        db.commit()
        return HTMLResponse("Forbidden — only an Owner can deactivate an Owner account.", status_code=403)
    if target.role == "owner" and target.active:
        other_active_owners = db.query(Staff).filter(
            Staff.role == "owner", Staff.active == True, Staff.id != target.id  # noqa: E712
        ).count()
        if other_active_owners == 0:
            return HTMLResponse("Forbidden — you can't deactivate the last active Owner account. "
                                 "Promote another staff member to Owner first.", status_code=403)

    target.active = not target.active
    add_audit(db, current, "STAFF_TOGGLE", f"{'Activated' if target.active else 'Deactivated'} {target.name}")
    db.commit()
    return RedirectResponse("/staff", status_code=303)


# ── SETTINGS ─────────────────────────────────────────────────────────
@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner"):
        return HTMLResponse("Forbidden — settings require owner role.", status_code=403)
    settings = {s.key: s.value for s in db.query(Setting).all()}
    flash = request.session.pop("flash", None)
    webhook_secret = get_sms_webhook_secret(db)
    sms_webhook_url = f"{str(request.base_url).rstrip('/')}/webhooks/sms-reply/{webhook_secret}"
    return templates.TemplateResponse(request, "settings.html", {
        "staff": staff, "settings": settings, "flash": flash,
        "security_question_set": bool(settings.get("security_answer_hash")),
        "sms_webhook_url": sms_webhook_url,
    })


@app.post("/settings")
def settings_save(request: Request, shop_name: str = Form(""), province: str = Form(""),
                   invoice_prefix: str = Form(""), shop_address: str = Form(""),
                   shop_phone: str = Form(""), shop_email: str = Form(""),
                   shop_gst: str = Form(""), shop_pst: str = Form(""),
                   points_per_dollar: float = Form(1), points_redeem_rate: float = Form(100),
                   email_method: str = Form("smtp"),
                   smtp_host: str = Form(""), smtp_port: str = Form(""), smtp_user: str = Form(""),
                   smtp_password: str = Form(""), smtp_from: str = Form(""),
                   brevo_api_key: str = Form(""), brevo_from_email: str = Form(""), brevo_from_name: str = Form(""),
                   twilio_sid: str = Form(""), twilio_token: str = Form(""), twilio_from: str = Form(""),
                   digest_enabled: str = Form(""), digest_email: str = Form(""), digest_hour: int = Form(21),
                   security_question: str = Form(""), security_answer: str = Form(""),
                   db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner"):
        return HTMLResponse("Forbidden", status_code=403)

    # These three are load-bearing (shop name on every receipt/label,
    # province drives tax calculation on every sale) — rather than
    # either 422-crashing the whole save on a framework validation
    # error, or silently persisting a blank shop name, catch it here
    # and bounce back with a clear message, leaving whatever was
    # already saved untouched.
    if not shop_name.strip() or not province.strip() or not invoice_prefix.strip():
        request.session["flash"] = ("red", "Shop Name, Province, and Invoice Prefix can't be blank — nothing was saved.")
        return RedirectResponse("/settings", status_code=303)

    set_setting(db, "shop_name", shop_name)
    set_setting(db, "province", province)
    set_setting(db, "invoice_prefix", invoice_prefix)
    set_setting(db, "shop_address", shop_address)
    set_setting(db, "shop_phone", shop_phone)
    set_setting(db, "shop_email", shop_email)
    set_setting(db, "shop_gst", shop_gst)
    set_setting(db, "shop_pst", shop_pst)
    set_setting(db, "points_per_dollar", str(points_per_dollar))
    set_setting(db, "points_redeem_rate", str(points_redeem_rate))
    set_setting(db, "email_method", "brevo_api" if email_method == "brevo_api" else "smtp")
    set_setting(db, "smtp_host", smtp_host)
    set_setting(db, "smtp_port", smtp_port)
    set_setting(db, "smtp_user", smtp_user)
    if smtp_password:  # only overwrite if a new one was actually typed
        set_setting(db, "smtp_password", encrypt_value(smtp_password))
    set_setting(db, "smtp_from", smtp_from)
    if brevo_api_key:  # only overwrite if a new one was actually typed
        set_setting(db, "brevo_api_key", encrypt_value(brevo_api_key))
    set_setting(db, "brevo_from_email", brevo_from_email)
    set_setting(db, "brevo_from_name", brevo_from_name)
    set_setting(db, "twilio_sid", twilio_sid)
    if twilio_token:
        set_setting(db, "twilio_token", encrypt_value(twilio_token))
    set_setting(db, "twilio_from", twilio_from)
    set_setting(db, "digest_enabled", "true" if digest_enabled == "on" else "false")
    set_setting(db, "digest_email", digest_email)
    set_setting(db, "digest_hour", str(digest_hour))
    if security_question.strip():
        set_setting(db, "security_question", security_question.strip())
    if security_answer.strip():  # only overwrite if a new answer was actually typed
        set_setting(db, "security_answer_hash", hash_pin(security_answer.strip().lower()))
    add_audit(db, staff, "SETTINGS", "Updated shop settings")
    _schedule_digest()  # pick up a new digest_hour immediately, not just on next restart
    return RedirectResponse("/settings", status_code=303)


# ── CASH SESSIONS ────────────────────────────────────────────────────
def todays_cash_card_totals(db: Session, today_str: str) -> dict:
    """The numbers Cash Up reconciles against physical reality: how much
    cash should be sitting in the drawer, and how much should show up on
    the card terminal's own batch report, for everything settled today.

    Two sources feed this, both filtered to today's date:
      - Invoice.cash_amount / card_amount — set at POS checkout, already
        split correctly for Cash/Card/Split sales (UPI, E-Transfer, and
        Store Credit contribute 0 to both, since none of those touch the
        drawer or the terminal).
      - LayawayPayment rows — each individual deposit/installment, on the
        day it was actually collected. The one big invoice created when a
        layaway is finally paid off is deliberately excluded here (see the
        comment at that Invoice(...) call) so an installment paid three
        weeks ago doesn't get counted again on payoff day.

    Used by both /cashup (to display today's running totals) and
    /cashup/close (to compute what a shift should have in the drawer) —
    previously each route had its own copy of this logic, string-matching
    payment_method values ("Credit Card", "Debit") that POS never actually
    wrote (it writes "Card"), so card reconciliation had never once
    matched anything.
    """
    today_invoices = [i for i in db.query(Invoice).all()
                       if i.date.strftime("%Y-%m-%d") == today_str and not i.refunded]
    cash_sales = sum(i.cash_amount for i in today_invoices)
    card_sales = sum(i.card_amount for i in today_invoices)
    etransfer_sales = sum(i.total for i in today_invoices if i.payment_method == "E-Transfer")
    total_sales = sum(i.total for i in today_invoices)

    today_layaway_payments = [p for p in db.query(LayawayPayment).all()
                               if p.created_at.strftime("%Y-%m-%d") == today_str]
    for p in today_layaway_payments:
        if p.method == "Cash":
            cash_sales += p.amount
        elif p.method in ("Debit", "Credit"):
            card_sales += p.amount
        # E-Transfer layaway installments settle themselves, same as a
        # regular E-Transfer sale — nothing to reconcile physically.

    card_invoices_today = [i for i in today_invoices if i.card_amount > 0]
    card_missing_ref = sum(1 for i in card_invoices_today if not i.card_reference)

    return {
        "cash_sales": round(cash_sales, 2), "card_sales": round(card_sales, 2),
        "etransfer_sales": round(etransfer_sales, 2), "total_sales": round(total_sales, 2),
        "invoice_count": len(today_invoices),
        "card_invoices_today": card_invoices_today, "card_missing_ref": card_missing_ref,
    }


@app.get("/cashup", response_class=HTMLResponse)
def cashup_page(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)

    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    totals = todays_cash_card_totals(db, today_str)

    cash_float = float(get_setting(db, "cash_float", "200"))
    expected = round(cash_float + totals["cash_sales"], 2)
    today_session = db.query(CashSession).filter(CashSession.date == today_str).first()
    history = db.query(CashSession).order_by(CashSession.date.desc()).limit(10).all()

    return templates.TemplateResponse(request, "cashup.html", {
        "staff": staff, "cash_float": cash_float, "expected": expected,
        "today_session": today_session, "history": history,
        **totals,
    })


@app.post("/cashup/close")
def cashup_close(request: Request, open_float: float = Form(...), actual: float = Form(...),
                  card_batch: str = Form(""), notes: str = Form(""), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)

    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    totals = todays_cash_card_totals(db, today_str)
    cash_sales = totals["cash_sales"]
    card_sales = totals["card_sales"]
    expected = round(open_float + cash_sales, 2)
    diff = round(actual - expected, 2)

    card_batch_val = float(card_batch) if card_batch.strip() else round(card_sales, 2)
    card_diff = round(card_batch_val - card_sales, 2)

    # Replace any existing session for today, same as the original
    existing = db.query(CashSession).filter(CashSession.date == today_str).first()
    if existing:
        db.delete(existing)
        db.flush()

    db.add(CashSession(
        date=today_str, open_float=open_float, expected=expected, actual=actual,
        difference=diff, card_expected=round(card_sales, 2), card_batch=card_batch_val,
        card_difference=card_diff, notes=notes, closed_by_id=staff.id, closed_by_name=staff.name,
    ))
    set_setting(db, "cash_float", str(open_float))
    audit_msg = f"Cash-up closed: cash expected ${expected:.2f}, actual ${actual:.2f}, diff ${diff:.2f}"
    if abs(card_diff) > 0.01:
        audit_msg += f" | card terminal mismatch ${card_diff:.2f}"
    add_audit(db, staff, "CASH_UP", audit_msg)
    db.commit()
    return RedirectResponse("/cashup", status_code=303)


def build_daily_digest(db: Session) -> tuple[str, str]:
    """Returns (subject, body) for the daily owner digest email."""
    today = datetime.utcnow().date()
    today_str = today.strftime("%Y-%m-%d")
    shop_name = get_setting(db, "shop_name", "Your Shop")

    today_invoices = [i for i in db.query(Invoice).all() if i.date.date() == today]
    paid = [i for i in today_invoices if not i.refunded]
    refunded = [i for i in today_invoices if i.refunded]
    revenue = round(sum(i.total for i in paid), 2)
    tax = round(sum(i.tax_total for i in paid), 2)

    repairs_opened = db.query(Repair).filter(Repair.created_at >= datetime.combine(today, datetime.min.time())).count()
    repairs_closed_today = sum(1 for r in db.query(Repair).filter(Repair.status.in_(["COMPLETED", "COLLECTED"])).all()
                                if r.updated_at and r.updated_at.date() == today)

    low_stock = db.query(Product).filter(Product.stock <= Product.reorder_threshold).all()

    cash_session = db.query(CashSession).filter(CashSession.date == today_str).first()

    lines = [
        f"Daily summary for {shop_name} — {today.strftime('%A, %B')} {today.day}, {today.year}",
        "",
        f"💰 Revenue: ${revenue:.2f} ({len(paid)} sale{'s' if len(paid) != 1 else ''})",
        f"🧾 Tax collected: ${tax:.2f}",
    ]
    if refunded:
        lines.append(f"↩️ Refunds today: {len(refunded)} (${sum(i.total for i in refunded):.2f})")
    lines += [
        "",
        f"🔧 Repair tickets opened today: {repairs_opened}",
        f"✅ Repair tickets closed today: {repairs_closed_today}",
        "",
    ]
    if low_stock:
        lines.append(f"📦 Low stock — {len(low_stock)} item(s) need reordering:")
        for p in low_stock[:10]:
            lines.append(f"   • {p.name} ({p.sku}) — {p.stock} left, reorder {p.reorder_qty}")
        if len(low_stock) > 10:
            lines.append(f"   ...and {len(low_stock) - 10} more. Full list: /products/reorder")
    else:
        lines.append("📦 Stock levels are fine — nothing needs reordering.")
    lines.append("")
    if cash_session:
        diff = cash_session.difference
        lines.append(f"🔒 Cash-up was closed today — difference: {'+'if diff>=0 else ''}{diff:.2f}")
    else:
        lines.append("🔒 Cash-up has not been closed yet today.")

    return f"{shop_name} — Daily Summary ({today_str})", "\n".join(lines)


@app.post("/settings/send-test-email")
def send_test_email(request: Request, to_email: str = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner"):
        return HTMLResponse("Forbidden", status_code=403)
    shop_name = get_setting(db, "shop_name", "TechPro+")
    method = get_setting(db, "email_method", "smtp")
    method_label = "Brevo's HTTPS API" if method == "brevo_api" else "SMTP"
    ok, msg = send_plain_email(
        db, to_email,
        f"Test email from {shop_name}",
        f"This is a test email from {shop_name}'s CRM. If you're reading this, {method_label} is working correctly.",
        get_setting,
    )
    request.session["flash"] = ("green" if ok else "red", msg)
    return RedirectResponse("/settings", status_code=303)


@app.get("/settings/smtp-diagnose", response_class=HTMLResponse)
def smtp_diagnose_page(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner"):
        return HTMLResponse("Forbidden", status_code=403)
    return templates.TemplateResponse(request, "smtp_diagnose.html", {
        "staff": staff, "steps": None, "to_email": "",
        "settings": {s.key: s.value for s in db.query(Setting).all()},
    })


@app.post("/settings/smtp-diagnose", response_class=HTMLResponse)
def smtp_diagnose_run(request: Request, to_email: str = Form(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    """Runs SMTP connection step-by-step and reports exactly which phase
    fails — lets you see the precise error rather than a generic message."""
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner"):
        return HTMLResponse("Forbidden", status_code=403)

    import socket as _socket

    host = get_setting(db, "smtp_host", "")
    port_str = get_setting(db, "smtp_port", "587")
    user = get_setting(db, "smtp_user", "")
    password = decrypt_value(get_setting(db, "smtp_password", ""))
    from_addr = get_setting(db, "smtp_from", "") or user
    shop_name = get_setting(db, "shop_name", "TechPro+")

    steps = []

    def s_ok(label, detail=""): steps.append(("ok", label, detail))
    def s_fail(label, detail=""): steps.append(("fail", label, detail))
    def s_warn(label, detail=""): steps.append(("warn", label, detail))

    ctx = {"settings": {s.key: s.value for s in db.query(Setting).all()},
           "staff": staff, "to_email": to_email, "steps": steps}

    def render(): return templates.TemplateResponse(request, "smtp_diagnose.html", ctx)

    # Step 1: settings completeness
    if not (host and port_str and user and password):
        s_fail("Settings check", "SMTP host, port, username, or password is missing in Settings.")
        return render()
    s_ok("Settings check", f"Host: {host}, Port: {port_str}, User: {user}, From: {from_addr}")

    # Step 2: TCP connectivity
    port = int(port_str)
    try:
        sock = _socket.create_connection((host, port), timeout=10)
        sock.close()
        s_ok("TCP connection", f"Reached {host}:{port}")
    except Exception as e:
        s_fail("TCP connection", f"Cannot connect to {host}:{port} — {e}. Your ISP or Windows firewall may be blocking outbound port {port}.")
        return render()

    # Step 3: TLS handshake
    import smtplib, ssl as _ssl
    server = None
    try:
        if port == 465:
            server = smtplib.SMTP_SSL(host, port, timeout=15, context=_ssl.create_default_context())
        else:
            server = smtplib.SMTP(host, port, timeout=15)
            server.ehlo()
            server.starttls(context=_ssl.create_default_context())
            server.ehlo()
        s_ok("TLS / STARTTLS", f"Encryption established")
    except Exception as e:
        s_fail("TLS / STARTTLS", str(e))
        return render()

    # Step 4: SMTP login
    try:
        server.login(user, password)
        s_ok("SMTP authentication", f"Logged in as {user}")
    except smtplib.SMTPAuthenticationError as e:
        s_fail("SMTP authentication",
               f"Credentials rejected — {e}. "
               f"For Brevo: the password must be your SMTP Key "
               f"(Brevo dashboard → SMTP & API → SMTP → Generate a new SMTP key). "
               f"It is NOT your Brevo account password.")
        server.quit()
        return render()
    except Exception as e:
        s_fail("SMTP authentication", str(e))
        server.quit()
        return render()

    # Step 5: Brevo-specific verified sender check
    is_brevo = "brevo" in host.lower() or "sendinblue" in host.lower()
    if is_brevo:
        s_warn("Brevo verified sender check",
               f"Brevo requires the From address '{from_addr}' to be added and verified in your "
               f"Brevo account before mail will actually be delivered. "
               f"Even if this test passes, if '{from_addr}' is not a verified sender, "
               f"Brevo accepts it at the SMTP level but silently drops it. "
               f"→ Go to: app.brevo.com → Senders & IPs → Senders → Add a sender.")
    else:
        s_ok("From address", f"Sending as {from_addr}")

    # Step 6: actually send the test message
    from email.mime.text import MIMEText
    from email.utils import formatdate, make_msgid
    try:
        msg = MIMEText(
            f"SMTP diagnostic test from {shop_name} CRM.\n\n"
            f"Connection details:\n"
            f"  Server: {host}:{port}\n"
            f"  Auth user: {user}\n"
            f"  From address: {from_addr}\n"
            f"  To: {to_email}\n\n"
            f"If you are reading this, all 6 steps passed successfully.\n\n"
            f"If you are using Brevo and this message did NOT arrive:\n"
            f"  → The From address '{from_addr}' is not yet verified in your Brevo account.\n"
            f"  → Go to app.brevo.com → Senders & IPs → Senders → Add a sender.\n"
            f"  → Then try sending a receipt again."
        )
        msg["Subject"] = f"[SMTP Diagnostic] {shop_name} CRM — step-by-step test"
        msg["From"] = from_addr
        msg["To"] = to_email
        msg["Date"] = formatdate(localtime=True)
        msg["Message-ID"] = make_msgid()
        refused = server.send_message(msg)
        server.quit()

        if refused:
            s_fail("Message accepted by server",
                   f"Server refused the recipient address: {refused}. "
                   f"Double-check the customer's email address.")
        else:
            s_ok("Message accepted by server",
                 f"✓ {host} accepted the message for delivery to {to_email}. "
                 f"Check inbox and spam folder. "
                 + ("If it doesn't arrive, the Brevo verified-sender issue above is most likely the cause."
                    if is_brevo else
                    "If it doesn't arrive within 2 minutes, check spam/junk."))
    except Exception as e:
        s_fail("Sending message", str(e))

    return render()


@app.post("/settings/send-test-digest")
def send_test_digest(request: Request, db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner"):
        return HTMLResponse("Forbidden", status_code=403)
    recipient = get_setting(db, "digest_email", "")
    if not recipient:
        request.session["flash"] = ("red", "Set a digest email address first.")
        return RedirectResponse("/settings", status_code=303)
    subject, body = build_daily_digest(db)
    ok, msg = send_plain_email(db, recipient, subject, body, get_setting)
    request.session["flash"] = ("green" if ok else "red", msg)
    return RedirectResponse("/settings", status_code=303)


def run_daily_digest_job():
    """Called by the scheduler — opens its own DB session since it runs
    outside any request."""
    db = SessionLocal()
    try:
        enabled = get_setting(db, "digest_enabled", "false") == "true"
        recipient = get_setting(db, "digest_email", "")
        if not (enabled and recipient):
            return
        subject, body = build_daily_digest(db)
        send_plain_email(db, recipient, subject, body, get_setting)
    finally:
        db.close()


_scheduler = BackgroundScheduler()
def _schedule_digest():
    db = SessionLocal()
    try:
        hour = int(get_setting(db, "digest_hour", "21"))
    finally:
        db.close()
    _scheduler.add_job(run_daily_digest_job, "cron", hour=hour, id="daily_digest", replace_existing=True)
_schedule_digest()
_scheduler.start()


# ── REPORTS ──────────────────────────────────────────────────────────
@app.get("/reports", response_class=HTMLResponse)
def reports_page(request: Request, tax_month: str = "", db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — reports require manager or owner role.", status_code=403)

    now = datetime.utcnow()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_end = month_start - timedelta(seconds=1)
    last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    all_invoices = db.query(Invoice).filter(Invoice.refunded == False).all()  # noqa: E712
    this_month = [i for i in all_invoices if i.date >= month_start]
    last_month = [i for i in all_invoices if last_month_start <= i.date <= last_month_end]

    month_revenue = round(sum(i.total for i in this_month), 2)
    last_month_revenue = round(sum(i.total for i in last_month), 2)
    month_tax = round(sum(i.tax_total for i in this_month), 2)
    revenue_trend_pct = round(((month_revenue - last_month_revenue) / last_month_revenue) * 100) if last_month_revenue else None

    # Profit: (price - product cost) * qty, across all (non-refunded) invoices
    profit = 0.0
    cogs = 0.0
    by_category = {}
    by_payment = {}
    prod_sales = {}
    staff_perf = {}
    for inv in all_invoices:
        pm_label = payment_method_label(inv.payment_method)
        by_payment[pm_label] = by_payment.get(pm_label, 0) + inv.total
        sname = inv.staff.name if inv.staff else "Unknown"
        sp = staff_perf.setdefault(sname, {"name": sname, "invoice_count": 0, "revenue": 0.0})
        sp["invoice_count"] += 1
        sp["revenue"] += inv.total
        for line in inv.lines:
            product = db.get(Product, line.product_id) if line.product_id else None
            line_total = line.price * line.qty
            if product:
                profit += (line.price - product.cost) * line.qty
                cogs += product.cost * line.qty
                cat = product.category or "Uncategorized"
                prod_sales.setdefault(line.product_id, {"name": line.name, "qty": 0, "rev": 0})
                prod_sales[line.product_id]["qty"] += line.qty
                prod_sales[line.product_id]["rev"] += line_total
            else:
                profit += line_total
                cat = "Repair / Service"
            by_category[cat] = by_category.get(cat, 0) + line_total

    top_products = sorted(prod_sales.values(), key=lambda p: p["qty"], reverse=True)[:8]
    staff_performance = sorted(staff_perf.values(), key=lambda s: s["revenue"], reverse=True)

    # Sales by day, last 14 days — for the Reports bar chart. Built the
    # same way the dashboard's sparkline already is (no charting
    # library — plain data the template turns into SVG bars).
    daily_totals = {}
    chart_start = (now - timedelta(days=13)).replace(hour=0, minute=0, second=0, microsecond=0)
    for inv in all_invoices:
        if inv.date >= chart_start:
            day_key = inv.date.strftime("%Y-%m-%d")
            daily_totals[day_key] = daily_totals.get(day_key, 0) + inv.total
    sales_by_day = []
    for i in range(14):
        day = chart_start + timedelta(days=i)
        key = day.strftime("%Y-%m-%d")
        # NOT using %-d here — it's Linux-only and crashes on Windows
        # (same reason the custom `datefmt` Jinja filter exists elsewhere).
        sales_by_day.append({"label": f"{day.strftime('%b')} {day.day}",
                              "total": round(daily_totals.get(key, 0), 2)})

    # Repairs closed per technician, this month — completes the staff performance picture
    repairs_this_month = db.query(Repair).filter(
        Repair.status.in_(["COMPLETED", "COLLECTED"]), Repair.updated_at >= month_start
    ).all()
    tech_repairs = {}
    for r in repairs_this_month:
        tname = r.technician.name if r.technician else "Unassigned"
        tech_repairs[tname] = tech_repairs.get(tname, 0) + 1

    # Tax summary for a selectable month (defaults to current month) — for CRA filing periods
    if not tax_month:
        tax_month = now.strftime("%Y-%m")
    ty, tm = int(tax_month[:4]), int(tax_month[5:7])
    tax_month_invoices = [i for i in all_invoices if i.date.year == ty and i.date.month == tm]
    tax_summary = {}
    for inv in tax_month_invoices:
        try:
            for line in json.loads(inv.tax_breakdown or "[]"):
                tax_summary[line["label"]] = tax_summary.get(line["label"], 0) + line["amount"]
        except (ValueError, KeyError, TypeError):
            continue
    tax_summary_total = round(sum(tax_summary.values()), 2)
    available_tax_months = sorted({i.date.strftime("%Y-%m") for i in all_invoices}, reverse=True)

    return templates.TemplateResponse(request, "reports.html", {
        "staff": staff, "month_revenue": month_revenue, "last_month_revenue": last_month_revenue,
        "revenue_trend_pct": revenue_trend_pct, "month_tax": month_tax,
        "month_invoice_count": len(this_month), "profit": round(profit, 2), "cogs": round(cogs, 2),
        "by_category": sorted(by_category.items(), key=lambda x: x[1], reverse=True),
        "by_payment": sorted(by_payment.items(), key=lambda x: x[1], reverse=True),
        "top_products": top_products, "staff_performance": staff_performance,
        "sales_by_day": sales_by_day,
        "tech_repairs": sorted(tech_repairs.items(), key=lambda x: x[1], reverse=True),
        "tax_month": tax_month, "tax_summary": sorted(tax_summary.items()),
        "tax_summary_total": tax_summary_total, "available_tax_months": available_tax_months,
    })


def _build_eod_data(db: Session, month_str: str):
    """Shared by both the on-screen EOD report and the printable version."""
    year, month = int(month_str[:4]), int(month_str[5:7])
    invoices = (db.query(Invoice)
                .filter(Invoice.refunded == False)  # noqa: E712
                .order_by(Invoice.date).all())
    days = {}
    for inv in invoices:
        if inv.date.year != year or inv.date.month != month:
            continue
        key = inv.date.strftime("%Y-%m-%d")
        d = days.setdefault(key, {
            "date": key, "weekday": inv.date.strftime("%A"),
            "count": 0, "subtotal": 0.0, "tax": 0.0, "total": 0.0,
            "by_payment": {}, "invoices": [],
        })
        d["count"] += 1
        d["subtotal"] += inv.subtotal
        d["tax"] += inv.tax_total
        d["total"] += inv.total
        pm_label = payment_method_label(inv.payment_method)
        d["by_payment"][pm_label] = d["by_payment"].get(pm_label, 0) + inv.total
        d["invoices"].append(inv)

    day_list = sorted(days.values(), key=lambda d: d["date"], reverse=True)
    grand_revenue = round(sum(d["total"] for d in day_list), 2)
    grand_tax = round(sum(d["tax"] for d in day_list), 2)
    grand_count = sum(d["count"] for d in day_list)
    daily_avg = round(grand_revenue / len(day_list), 2) if day_list else 0
    return day_list, grand_revenue, grand_tax, grand_count, daily_avg


@app.get("/reports/eod", response_class=HTMLResponse)
def eod_report(request: Request, month: str = "", db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — reports require manager or owner role.", status_code=403)
    month = month or datetime.utcnow().strftime("%Y-%m")
    day_list, grand_revenue, grand_tax, grand_count, daily_avg = _build_eod_data(db, month)
    return templates.TemplateResponse(request, "eod_report.html", {
        "staff": staff, "month": month, "days": day_list,
        "grand_revenue": grand_revenue, "grand_tax": grand_tax,
        "grand_count": grand_count, "daily_avg": daily_avg,
    })


@app.get("/reports/eod/print", response_class=HTMLResponse)
def eod_report_print(request: Request, month: str = "", day: str = "", db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — reports require manager or owner role.", status_code=403)
    month = month or datetime.utcnow().strftime("%Y-%m")
    day_list, grand_revenue, grand_tax, grand_count, daily_avg = _build_eod_data(db, month)
    if day:
        day_list = [d for d in day_list if d["date"] == day]
        grand_revenue = round(sum(d["total"] for d in day_list), 2)
        grand_tax = round(sum(d["tax"] for d in day_list), 2)
        grand_count = sum(d["count"] for d in day_list)
    shop_name = get_setting(db, "shop_name", "TechPro+")
    return templates.TemplateResponse(request, "eod_print.html", {
        "staff": staff, "month": month, "single_day": day, "days": day_list,
        "grand_revenue": grand_revenue, "grand_tax": grand_tax,
        "grand_count": grand_count, "shop_name": shop_name,
    })


# ── DATA EXPORT / IMPORT / BACKUP ───────────────────────────────────
def _csv_response(rows, headers, filename):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return Response(content=buf.getvalue(), media_type="text/csv",
                     headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/export/backup")
def export_backup(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner"):
        return HTMLResponse("Forbidden — backups require owner role.", status_code=403)

    def row(obj, exclude=()):
        return {c.name: getattr(obj, c.name) for c in obj.__table__.columns if c.name not in exclude}

    data = {
        "version": 1, "exported_at": datetime.utcnow().isoformat(),
        "settings": {s.key: s.value for s in db.query(Setting).all()},
        "products": [row(p) for p in db.query(Product).all()],
        "customers": [row(c) for c in db.query(Customer).all()],
        "repairs": [row(r) for r in db.query(Repair).all()],
        "invoices": [dict(row(i), lines=[row(l) for l in i.lines]) for i in db.query(Invoice).all()],
        # Staff PIN hashes are intentionally excluded from the backup file —
        # a leaked backup shouldn't double as a leaked set of login credentials.
        # Staff records (names/roles) are kept; PINs must be re-set after a restore.
        "staff": [row(s, exclude=("pin_hash",)) for s in db.query(Staff).all()],
    }
    set_setting(db, "last_backup", datetime.utcnow().isoformat())
    add_audit(db, staff, "BACKUP_EXPORT", "Full data backup exported")
    filename = f"techpro_backup_{datetime.utcnow().strftime('%Y-%m-%d')}.json"
    return Response(content=json.dumps(data, indent=2, default=str), media_type="application/json",
                     headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/import/backup")
async def import_backup(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db), _csrf: None = Depends(csrf_protect)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner"):
        return HTMLResponse("Forbidden — restoring backups requires owner role.", status_code=403)

    try:
        data = json.loads(await file.read())
    except Exception as e:
        return HTMLResponse(f"Could not parse backup file: {e}", status_code=400)
    if not data.get("version"):
        return HTMLResponse("This doesn't look like a valid backup file.", status_code=400)

    try:
        if "settings" in data:
            for k, v in data["settings"].items():
                set_setting(db, k, v)
        if "products" in data:
            db.query(Product).delete()
            for p in data["products"]:
                db.add(Product(**p))  # keep original id so existing invoice line references stay valid
        if "customers" in data:
            db.query(Customer).delete()
            for c in data["customers"]:
                db.add(Customer(**c))
        if "repairs" in data:
            db.query(Repair).delete()
            for r in data["repairs"]:
                for dt_field in ("created_at", "updated_at"):
                    if r.get(dt_field) and isinstance(r[dt_field], str):
                        try:
                            r[dt_field] = datetime.fromisoformat(r[dt_field])
                        except ValueError:
                            r[dt_field] = datetime.utcnow()
                db.add(Repair(**r))
        # Invoices/lines intentionally left alone on restore — overwriting sales
        # history is rarely what you want from a "restore my catalog/customers"
        # action. Flag this clearly to the person rather than silently doing it.
    except (TypeError, ValueError) as e:
        db.rollback()
        add_audit(db, staff, "SECURITY_BLOCK", f"Backup restore aborted — malformed file: {e}")
        db.commit()
        return HTMLResponse(
            f"Restore aborted — the backup file doesn't match this app's data format ({e}). "
            f"Nothing was changed; your existing data is untouched.", status_code=400)

    add_audit(db, staff, "BACKUP_IMPORT", "Backup restored (products/customers/repairs/settings)")
    db.commit()
    return RedirectResponse("/settings", status_code=303)


@app.get("/export/csv/inventory")
def export_inventory_csv(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — exports require manager or owner role.", status_code=403)
    rows = [(p.sku, p.name, p.category, p.subcategory, p.price, p.cost, p.stock) for p in db.query(Product).all()]
    return _csv_response(rows, ["SKU", "Name", "Category", "Brand", "Price", "Cost", "Stock"], "inventory.csv")


@app.get("/export/csv/customers")
def export_customers_csv(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — exports require manager or owner role.", status_code=403)
    rows = [(c.name, c.phone, c.email, c.spent, c.points, c.store_credit) for c in db.query(Customer).all()]
    return _csv_response(rows, ["Name", "Phone", "Email", "Total Spent", "Points", "Store Credit"], "customers.csv")


@app.get("/export/csv/invoices")
def export_invoices_csv(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — exports require manager or owner role.", status_code=403)
    rows = [(i.number, i.customer.name if i.customer else "Walk-in", i.payment_method,
             i.subtotal, i.discount, i.tax_total, i.total, i.date.strftime("%Y-%m-%d %H:%M"),
             "Refunded" if i.refunded else "Paid") for i in db.query(Invoice).order_by(Invoice.date).all()]
    return _csv_response(rows, ["Number", "Customer", "Payment", "Subtotal", "Discount", "Tax", "Total", "Date", "Status"], "invoices.csv")


@app.get("/export/csv/tax-report")
def export_tax_report_csv(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — exports require manager or owner role.", status_code=403)
    rows = [(i.number, i.date.strftime("%Y-%m-%d"), i.subtotal, i.discount, i.tax_total, i.total)
            for i in db.query(Invoice).filter(Invoice.refunded == False).order_by(Invoice.date).all()]  # noqa: E712
    return _csv_response(rows, ["Invoice", "Date", "Subtotal", "Discount", "Tax Collected", "Total"], "tax_report.csv")


# Bold, slightly larger header rows for every export sheet — small touch,
# but a workbook that gets handed to an accountant/bookkeeper reads a lot
# better with headers that are visually distinct from data rows.
_XLSX_HEADER_FONT = Font(bold=True)


def _autosize_columns(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_header(ws, headers):
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = _XLSX_HEADER_FONT


@app.get("/export/xlsx/month-end")
def export_month_end_xlsx(request: Request, month: str = "", db: Session = Depends(get_db)):
    """Month-end workbook for handing off to a bookkeeper/accountant:
    every invoice for the selected month with its tax breakdown, a top
    products sheet, and a tax-summary-by-rate sheet lined up with the
    same CRA filing-period logic the Reports page already uses. Defaults
    to the current month; pass ?month=YYYY-MM for any other period."""
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — exports require manager or owner role.", status_code=403)

    now = datetime.utcnow()
    if not month:
        month = now.strftime("%Y-%m")
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        return HTMLResponse("Invalid month — expected format YYYY-MM.", status_code=400)

    invoices = (
        db.query(Invoice)
        .filter(Invoice.date >= datetime(year, mon, 1))
        .filter(Invoice.date < (datetime(year + 1, 1, 1) if mon == 12 else datetime(year, mon + 1, 1)))
        .order_by(Invoice.date)
        .all()
    )

    wb = Workbook()

    # ── Sheet 1: Invoices ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Invoices"
    _write_header(ws1, ["Invoice #", "Date", "Customer", "Payment Method",
                         "Subtotal", "Discount", "Tax Detail", "Tax Total", "Total", "Status"])
    for inv in invoices:
        try:
            tax_lines = json.loads(inv.tax_breakdown or "[]")
            tax_detail = "; ".join(f"{t['label']}: ${t['amount']:.2f}" for t in tax_lines)
        except (ValueError, KeyError, TypeError):
            tax_detail = ""
        ws1.append([
            inv.number,
            inv.date.strftime("%Y-%m-%d %H:%M"),
            inv.customer.name if inv.customer else "Walk-in",
            inv.payment_method,
            float(inv.subtotal),
            float(inv.discount),
            tax_detail,
            float(inv.tax_total),
            float(inv.total),
            "Refunded" if inv.refunded else "Paid",
        ])
    _autosize_columns(ws1, [14, 18, 20, 16, 12, 12, 34, 12, 12, 12])
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Top Products (by qty sold, this month) ─────────────
    prod_sales = {}
    for inv in invoices:
        if inv.refunded:
            continue
        for line in inv.lines:
            key = line.product_id or line.name
            entry = prod_sales.setdefault(key, {"name": line.name, "qty": 0, "revenue": 0.0})
            entry["qty"] += line.qty
            entry["revenue"] += line.price * line.qty
    top_products = sorted(prod_sales.values(), key=lambda p: p["qty"], reverse=True)

    ws2 = wb.create_sheet("Top Products")
    _write_header(ws2, ["Product", "Qty Sold", "Revenue"])
    for p in top_products:
        ws2.append([p["name"], p["qty"], round(p["revenue"], 2)])
    _autosize_columns(ws2, [34, 12, 14])
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Tax Summary (by rate/label, non-refunded only) ─────
    tax_summary = {}
    for inv in invoices:
        if inv.refunded:
            continue
        try:
            for line in json.loads(inv.tax_breakdown or "[]"):
                tax_summary[line["label"]] = tax_summary.get(line["label"], 0) + line["amount"]
        except (ValueError, KeyError, TypeError):
            continue

    ws3 = wb.create_sheet("Tax Summary")
    _write_header(ws3, ["Tax", "Amount Collected"])
    for label, amount in sorted(tax_summary.items()):
        ws3.append([label, round(amount, 2)])
    ws3.append(["Total", round(sum(tax_summary.values()), 2)])
    ws3[ws3.max_row][0].font = _XLSX_HEADER_FONT
    ws3[ws3.max_row][1].font = _XLSX_HEADER_FONT
    _autosize_columns(ws3, [24, 18])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    add_audit(db, staff, "EXPORT", f"Month-end workbook exported for {month}")
    db.commit()

    filename = f"month_end_{month}.xlsx"
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── AUDIT LOG ────────────────────────────────────────────────────────
@app.get("/audit", response_class=HTMLResponse)
def audit_log(request: Request, db: Session = Depends(get_db)):
    staff = require_login(request, db)
    if not staff:
        return RedirectResponse("/login", status_code=303)
    if not role_allowed(staff, "owner", "manager"):
        return HTMLResponse("Forbidden — audit log requires manager or owner role.", status_code=403)
    logs = db.query(AuditLog).order_by(AuditLog.ts.desc()).limit(200).all()
    return templates.TemplateResponse(request, "audit.html", {"staff": staff, "logs": logs})
